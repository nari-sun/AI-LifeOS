import base64
import importlib.util
import io
import json
import os
import subprocess
import sys
import tempfile
import time
import unittest
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest import mock

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

import chat_gui_bridge  # noqa: E402
import build_answer_context  # noqa: E402
import codex_conversation  # noqa: E402
import kokoro_tts  # noqa: E402
from live_session import create_live_session  # noqa: E402


class ChatGuiBridgeTests(unittest.TestCase):
    def test_read_aloud_passes_only_temporary_audio_metadata_and_does_not_log_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            audio_path = root / "temporary.wav"
            audio_path.write_bytes(b"RIFF")
            private_text = "PRIVATE_READ_ALOUD_TEXT"

            with mock.patch.object(chat_gui_bridge, "synthesize_to_wav", return_value=audio_path) as synthesize:
                result = chat_gui_bridge.handle_read_aloud(
                    {
                        "root": str(root),
                        "request_id": "tts-test",
                        "voice": "jf_alpha",
                        "text": private_text,
                    }
                )

            self.assertEqual("tts-test", result["request_id"])
            self.assertEqual("jf_alpha", result["voice"])
            self.assertEqual(str(audio_path.resolve()), result["audio_path"])
            self.assertEqual(private_text, synthesize.call_args.kwargs["text"])
            self.assertEqual("jf_alpha", synthesize.call_args.kwargs["voice"])
            log_text = (root / "logs" / "tts" / "kokoro_tts.log").read_text(encoding="utf-8")
            self.assertNotIn(private_text, log_text)

    def test_read_aloud_rejects_voice_outside_allowlist(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "voice"):
                chat_gui_bridge.handle_read_aloud(
                    {
                        "root": temp_dir,
                        "request_id": "tts-test",
                        "voice": "unknown",
                        "text": "テスト",
                    }
                )

    def test_read_aloud_stream_publishes_each_audio_chunk_without_logging_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first.wav"
            second = root / "second.wav"
            first.write_bytes(b"RIFF")
            second.write_bytes(b"RIFF")
            events = []

            def fake_synthesize(**kwargs):
                kwargs["on_chunk"](first, 0)
                kwargs["on_chunk"](second, 1)
                return [first, second]

            with mock.patch.object(chat_gui_bridge, "synthesize_to_wav_chunks", side_effect=fake_synthesize):
                result = chat_gui_bridge.handle_read_aloud_stream(
                    {
                        "root": str(root),
                        "request_id": "tts-stream",
                        "voice": "jf_alpha",
                        "text": "PRIVATE_STREAM_TEXT",
                    },
                    events.append,
                )

            self.assertEqual(2, result["chunk_count"])
            self.assertEqual([0, 1], [event["index"] for event in events])
            self.assertEqual([str(first.resolve()), str(second.resolve())], [event["audio_path"] for event in events])
            log_text = (root / "logs" / "tts" / "kokoro_tts.log").read_text(encoding="utf-8")
            self.assertNotIn("PRIVATE_STREAM_TEXT", log_text)

    def test_read_aloud_honors_a_stop_request_that_arrives_before_synthesis_starts(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runtime_dir = root / "runtime-tts"
            runtime_dir.mkdir()
            cancel_file = runtime_dir / "read-aloud-tts-test.cancel"
            cancel_file.write_text("cancel\n", encoding="utf-8")

            with mock.patch.object(chat_gui_bridge, "_read_aloud_runtime_dir", return_value=runtime_dir):
                with self.assertRaisesRegex(kokoro_tts.KokoroSynthesisCancelled, "停止"):
                    chat_gui_bridge.handle_read_aloud(
                        {
                            "root": str(root),
                            "request_id": "tts-test",
                            "voice": "jf_alpha",
                            "text": "テスト",
                        }
                    )

            self.assertFalse(cancel_file.exists())

    def test_cancel_and_discard_read_aloud_are_limited_to_runtime_temp_audio(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runtime_dir = root / "runtime-tts"
            runtime_dir.mkdir()
            audio_path = runtime_dir / "read-aloud-tts-test.wav"
            audio_path.write_bytes(b"RIFF")

            with mock.patch.object(chat_gui_bridge, "_read_aloud_runtime_dir", return_value=runtime_dir):
                cancelled = chat_gui_bridge.handle_cancel_read_aloud({"root": str(root), "request_id": "tts-test"})
                discarded = chat_gui_bridge.handle_discard_read_aloud_audio({"root": str(root), "audio_path": str(audio_path)})

            self.assertTrue(cancelled["cancelled"])
            self.assertTrue(discarded["removed"])
            self.assertFalse(audio_path.exists())
            with self.assertRaisesRegex(ValueError, "一時音声"):
                chat_gui_bridge.handle_discard_read_aloud_audio({"root": str(root), "audio_path": str(root / "outside.wav")})

    def test_start_session_returns_live_jsonl_path(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)

            result = chat_gui_bridge.handle_start_session({"root": str(root)})

            self.assertEqual(result["messages"], [])
            self.assertTrue(result["session"]["jsonl_file"].startswith("inbox"))
            self.assertTrue((root / result["session"]["jsonl_file"]).parent.exists())

    def test_chatgpt_import_preview_and_apply_require_explicit_selection(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "lifeos"
            source_dir = Path(temp_dir) / "chatgpt-export"
            source_dir.mkdir()
            source_path = source_dir / "conversations.json"
            source_path.write_text(
                json.dumps(
                    [
                        {
                            "id": "gui-import-1",
                            "title": "GUI import conversation",
                            "create_time": 1767225600,
                            "update_time": 1767225660,
                            "current_node": "assistant-node",
                            "mapping": {
                                "root": {"id": "root", "parent": None, "message": None},
                                "user-node": {
                                    "id": "user-node",
                                    "parent": "root",
                                    "message": {
                                        "author": {"role": "user"},
                                        "create_time": 1767225601,
                                        "content": {
                                            "parts": [
                                                "GUIから取り込みたい",
                                                {"content_type": "audio_transcription", "text": "音声からの発言"},
                                                {"content_type": "image_asset_pointer"},
                                            ]
                                        },
                                    },
                                },
                                "assistant-node": {
                                    "id": "assistant-node",
                                    "parent": "user-node",
                                    "message": {
                                        "author": {"role": "assistant"},
                                        "create_time": 1767225602,
                                        "content": {"parts": ["安全に確認します"]},
                                    },
                                },
                            },
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            preview = chat_gui_bridge.handle_preview_chatgpt_import({"root": str(root), "source": str(source_path)})

            self.assertEqual(1, preview["total_count"])
            self.assertEqual(1, preview["new_count"])
            self.assertEqual(0, preview["updated_count"])
            self.assertEqual(0, preview["conflict_count"])
            self.assertFalse(preview["conversations"][0]["duplicate"])
            self.assertEqual("new", preview["conversations"][0]["import_state"])
            self.assertEqual(2, preview["conversations"][0]["source_message_count"])
            self.assertEqual(1, preview["conversations"][0]["attachment_count"])
            self.assertEqual(1, preview["conversations"][0]["audio_transcription_count"])
            self.assertFalse(preview["conversations"][0]["empty_conversation"])
            self.assertFalse((root / "conversations").exists())
            with self.assertRaisesRegex(ValueError, "最終確認"):
                chat_gui_bridge.handle_apply_chatgpt_import(
                    {"root": str(root), "source": str(source_path), "selected_ids": ["gui-import-1"]}
                )

            result = chat_gui_bridge.handle_apply_chatgpt_import(
                {
                    "root": str(root),
                    "source": str(source_path),
                    "selected_ids": ["gui-import-1"],
                    "confirmed": True,
                }
            )

            self.assertEqual(1, result["selected_count"])
            self.assertEqual(1, result["imported_count"])
            self.assertTrue(result["index_updated"])
            self.assertEqual("fresh", result["index_status"])
            self.assertIsNone(result["index_error"])
            raw_file = root / result["imported"][0]["raw_file"]
            self.assertTrue(raw_file.exists())
            self.assertTrue((raw_file.parent / "import_metadata.json").exists())
            self.assertFalse((root / "journal").exists())
            self.assertTrue((root / "memory" / "search_index.sqlite3").exists())
            refreshed_preview = chat_gui_bridge.handle_preview_chatgpt_import({"root": str(root), "source": str(source_path)})
            self.assertTrue(refreshed_preview["conversations"][0]["duplicate"])
            self.assertEqual("duplicate", refreshed_preview["conversations"][0]["import_state"])
            log_text = (root / "logs" / "chat_gui_bridge.log").read_text(encoding="utf-8")
            self.assertNotIn(str(source_path), log_text)

    def test_chatgpt_import_updates_changed_revision_and_rebuilds_index(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "lifeos"
            source_path = Path(temp_dir) / "conversations.json"

            def write_revision(update_time: int, assistant_text: str) -> None:
                source_path.write_text(
                    json.dumps(
                        [
                            {
                                "id": "gui-import-updated",
                                "title": "Updated export conversation",
                                "create_time": 1767225600,
                                "update_time": update_time,
                                "current_node": "assistant-node",
                                "mapping": {
                                    "user-node": {
                                        "id": "user-node",
                                        "parent": None,
                                        "message": {
                                            "author": {"role": "user"},
                                            "create_time": 1767225601,
                                            "content": {"parts": ["更新確認"]},
                                        },
                                    },
                                    "assistant-node": {
                                        "id": "assistant-node",
                                        "parent": "user-node",
                                        "message": {
                                            "author": {"role": "assistant"},
                                            "create_time": 1767225602,
                                            "content": {"parts": [assistant_text]},
                                        },
                                    },
                                },
                            }
                        ],
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )

            write_revision(1767225660, "最初の回答")
            first = chat_gui_bridge.handle_apply_chatgpt_import(
                {
                    "root": str(root),
                    "source": str(source_path),
                    "selected_ids": ["gui-import-updated"],
                    "confirmed": True,
                }
            )
            raw_file = root / first["imported"][0]["raw_file"]

            write_revision(1767225720, "更新後の回答")
            preview = chat_gui_bridge.handle_preview_chatgpt_import(
                {"root": str(root), "source": str(source_path)}
            )
            self.assertEqual(0, preview["new_count"])
            self.assertEqual(1, preview["updated_count"])
            self.assertEqual("updated", preview["conversations"][0]["import_state"])
            self.assertFalse(preview["conversations"][0]["duplicate"])

            updated = chat_gui_bridge.handle_apply_chatgpt_import(
                {
                    "root": str(root),
                    "source": str(source_path),
                    "selected_ids": ["gui-import-updated"],
                    "confirmed": True,
                }
            )

            self.assertEqual(0, updated["imported_count"])
            self.assertEqual(1, updated["updated_count"])
            self.assertEqual(0, updated["duplicate_count"])
            self.assertTrue(updated["index_updated"])
            self.assertTrue(updated["imported"][0]["updated"])
            self.assertEqual(raw_file, root / updated["imported"][0]["raw_file"])
            self.assertIn("更新後の回答", raw_file.read_text(encoding="utf-8"))
            refreshed = chat_gui_bridge.handle_preview_chatgpt_import(
                {"root": str(root), "source": str(source_path)}
            )
            self.assertEqual("duplicate", refreshed["conversations"][0]["import_state"])

    def test_chatgpt_import_conflict_is_previewed_but_not_applied(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "lifeos"
            source_path = Path(temp_dir) / "conversations.json"
            source_path.write_text(
                json.dumps(
                    [
                        {
                            "id": "gui-import-conflict",
                            "title": "First revision",
                            "create_time": 1767225600,
                            "update_time": 1767225660,
                            "mapping": {},
                        },
                        {
                            "id": "gui-import-conflict",
                            "title": "Second revision",
                            "create_time": 1767225600,
                            "update_time": 1767225720,
                            "mapping": {},
                        },
                    ]
                ),
                encoding="utf-8",
            )

            preview = chat_gui_bridge.handle_preview_chatgpt_import(
                {"root": str(root), "source": str(source_path)}
            )

            self.assertEqual(2, preview["conflict_count"])
            self.assertTrue(
                all(item["import_state"] == "conflict" for item in preview["conversations"])
            )
            with self.assertRaisesRegex(ValueError, "競合"):
                chat_gui_bridge.handle_apply_chatgpt_import(
                    {
                        "root": str(root),
                        "source": str(source_path),
                        "selected_ids": ["gui-import-conflict"],
                        "confirmed": True,
                    }
                )
            self.assertFalse((root / "conversations").exists())

    def test_chatgpt_import_keeps_raw_success_when_index_rebuild_fails(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "lifeos"
            source_path = Path(temp_dir) / "conversations.json"
            source_path.write_text(
                json.dumps(
                    [
                        {
                            "id": "gui-import-index-failure",
                            "title": "Private import title",
                            "create_time": 1767225600,
                            "mapping": {
                                "user-node": {
                                    "id": "user-node",
                                    "parent": None,
                                    "message": {
                                        "author": {"role": "user"},
                                        "create_time": 1767225601,
                                        "content": {"parts": ["PRIVATE_IMPORT_BODY"]},
                                    },
                                }
                            },
                        }
                    ]
                ),
                encoding="utf-8",
            )
            private_error = f"PRIVATE_INDEX_ERROR {source_path.resolve()} PRIVATE_IMPORT_BODY"

            with mock.patch.object(chat_gui_bridge, "rebuild_index", side_effect=RuntimeError(private_error)):
                result = chat_gui_bridge.handle_apply_chatgpt_import(
                    {
                        "root": str(root),
                        "source": str(source_path),
                        "selected_ids": ["gui-import-index-failure"],
                        "confirmed": True,
                    }
                )

            self.assertEqual(1, result["imported_count"])
            self.assertFalse(result["index_updated"])
            self.assertEqual("error", result["index_status"])
            self.assertIn("取り込みは完了", result["index_error"])
            self.assertNotIn("PRIVATE_INDEX_ERROR", result["index_error"])
            raw_file = root / result["imported"][0]["raw_file"]
            self.assertTrue(raw_file.exists())
            log_text = (root / "logs" / "chat_gui_bridge.log").read_text(encoding="utf-8")
            self.assertNotIn("PRIVATE_INDEX_ERROR", log_text)
            self.assertNotIn("PRIVATE_IMPORT_BODY", log_text)
            self.assertNotIn(str(source_path.resolve()), log_text)

    def test_chatgpt_import_preview_accepts_split_conversation_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "lifeos"
            source_path = Path(temp_dir) / "conversations-000.json"
            source_path.write_text("[]", encoding="utf-8")

            preview = chat_gui_bridge.handle_preview_chatgpt_import(
                {"root": str(root), "source": str(source_path)}
            )

            self.assertEqual(0, preview["total_count"])

    def test_send_message_saves_user_message_without_ai(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)

            result = chat_gui_bridge.handle_send_message(
                {
                    "root": str(root),
                    "content": "hello",
                    "no_ai": True,
                }
            )

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]

            self.assertEqual(records[0]["role"], "user")
            self.assertEqual(records[0]["content"], "hello")
            self.assertIsNone(result["assistant"])
            self.assertTrue(session_path.with_suffix(".session.json").exists())
            log_text = (root / "logs" / "chat_gui_bridge.log").read_text(encoding="utf-8")
            self.assertIn("send_message.done", log_text)
            self.assertNotIn("hello", log_text)

    def test_send_message_accepts_new_session_path_from_start_session(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            started = chat_gui_bridge.handle_start_session({"root": str(root)})
            session_file = started["session"]["jsonl_file"]

            result = chat_gui_bridge.handle_send_message(
                {
                    "root": str(root),
                    "session_file": session_file,
                    "content": "first message",
                    "no_ai": True,
                }
            )

            session_path = root / result["session"]["jsonl_file"]
            self.assertTrue(session_path.exists())
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]
            self.assertEqual(records[0]["content"], "first message")

    def test_send_message_forwards_full_archive_review_without_changing_saved_content(self):
        captured = {}

        def fake_generate(**kwargs):
            captured.update(kwargs)
            return codex_conversation.AssistantReplyResult(reply="確認しました。", memory_context=None)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            with mock.patch.object(chat_gui_bridge, "generate_assistant_reply_with_context", side_effect=fake_generate):
                result = chat_gui_bridge.handle_send_message(
                    {
                        "root": str(root),
                        "content": "傾向を教えて",
                        "full_archive_review": True,
                    }
                )

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]
            self.assertTrue(captured["force_full_archive_review"])
            self.assertEqual("傾向を教えて", records[0]["content"])

    def test_send_message_rejects_non_boolean_full_archive_review(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            with self.assertRaisesRegex(ValueError, "full_archive_review"):
                chat_gui_bridge.handle_send_message(
                    {
                        "root": str(root),
                        "content": "hello",
                        "full_archive_review": "true",
                    }
                )

            self.assertFalse((root / "inbox" / "live").exists())

    def test_streaming_send_emits_deltas_and_saves_only_completed_reply(self):
        original = chat_gui_bridge.generate_assistant_reply_streaming_with_context
        captured = {}

        def fake_stream(root, messages, on_delta, **kwargs):
            captured.update(kwargs)
            on_delta("途中")
            on_delta("返答")
            return codex_conversation.AssistantReplyResult(reply="確定返答", memory_context=None)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            deltas = []
            chat_gui_bridge.generate_assistant_reply_streaming_with_context = fake_stream
            try:
                result = chat_gui_bridge.handle_send_message(
                    {
                        "root": str(root),
                        "content": "hello",
                        "request_id": "stream-test",
                        "full_archive_review": True,
                    },
                    on_delta=deltas.append,
                )
            finally:
                chat_gui_bridge.generate_assistant_reply_streaming_with_context = original

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]
            self.assertEqual(["途中", "返答"], deltas)
            self.assertTrue(captured["force_full_archive_review"])
            self.assertEqual(["user", "assistant"], [record["role"] for record in records])
            self.assertEqual("確定返答", records[1]["content"])

    def test_streaming_cancel_discards_partial_reply(self):
        original = chat_gui_bridge.generate_assistant_reply_streaming_with_context

        def fake_stream(root, messages, on_delta, **kwargs):
            on_delta("保存しない途中返答")
            raise InterruptedError("返答生成を停止しました。")

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            deltas = []
            chat_gui_bridge.generate_assistant_reply_streaming_with_context = fake_stream
            try:
                result = chat_gui_bridge.handle_send_message(
                    {"root": str(root), "content": "hello", "request_id": "cancel-test"},
                    on_delta=deltas.append,
                )
            finally:
                chat_gui_bridge.generate_assistant_reply_streaming_with_context = original

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]
            self.assertEqual(["保存しない途中返答"], deltas)
            self.assertTrue(result["cancelled"])
            self.assertEqual(["user"], [record["role"] for record in records])

    def test_streaming_unavailable_falls_back_without_duplicate_user_message(self):
        original_stream = chat_gui_bridge.generate_assistant_reply_streaming_with_context
        original_generate = chat_gui_bridge.generate_assistant_reply_with_context

        def unavailable(**kwargs):
            raise codex_conversation.AppServerStreamingUnavailable("unsupported")

        def fake_generate(**kwargs):
            return codex_conversation.AssistantReplyResult(reply="fallback", memory_context=None)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            chat_gui_bridge.generate_assistant_reply_streaming_with_context = unavailable
            chat_gui_bridge.generate_assistant_reply_with_context = fake_generate
            try:
                result = chat_gui_bridge.handle_send_message(
                    {"root": str(root), "content": "hello", "request_id": "fallback-test"},
                    on_delta=lambda _: None,
                )
            finally:
                chat_gui_bridge.generate_assistant_reply_streaming_with_context = original_stream
                chat_gui_bridge.generate_assistant_reply_with_context = original_generate

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]
            self.assertEqual(["user", "assistant"], [record["role"] for record in records])
            self.assertEqual("fallback", records[1]["content"])

    def test_send_message_returns_memory_context_metadata(self):
        original_generate = chat_gui_bridge.generate_assistant_reply_with_context

        def fake_generate(root, messages, **kwargs):
            context = build_answer_context.build_answer_context(
                root=root,
                question=messages[-1].content,
                use_index=False,
            )
            return codex_conversation.AssistantReplyResult(reply="記憶を参照した返答です。", memory_context=context)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            memory = root / "memory"
            memory.mkdir()
            (memory / "long_term.md").write_text("# Long-Term Memory\n\n- ユーザーはAI-LifeOSを作っている。\n", encoding="utf-8")
            (memory / "preferences.md").write_text("# Preferences\n\n- ユーザーは静かな店を好む。\n", encoding="utf-8")

            chat_gui_bridge.generate_assistant_reply_with_context = fake_generate
            try:
                result = chat_gui_bridge.handle_send_message(
                    {
                        "root": str(root),
                        "content": "俺の好みに合う店は？",
                    }
                )
            finally:
                chat_gui_bridge.generate_assistant_reply_with_context = original_generate

            self.assertEqual("assistant", result["assistant"]["role"])
            self.assertTrue(result["memory_context"]["used"])
            self.assertGreaterEqual(result["memory_context"]["reference_count"], 1)
            paths = {reference["path"] for reference in result["memory_context"]["references"]}
            self.assertIn("memory/preferences.md", paths)

    def test_memory_context_reference_serializes_speaker_role(self):
        reference = build_answer_context.MemoryContextReference(
            path="conversations/2026/07/example/raw.md",
            document_type="raw_chunk",
            title="Example / Assistant message 2",
            date="2026-07-11",
            snippet="ROLE_ASSISTANT_SENTINEL",
            score=4,
            speaker_role="assistant",
            message_number=2,
        )

        result = chat_gui_bridge._serialize_memory_reference(reference)

        self.assertEqual("assistant", result["speaker_role"])
        self.assertEqual(2, result["message_number"])

    def test_memory_context_serializes_retrieval_modes(self):
        context = build_answer_context.AnswerContext(
            should_use_memory=True,
            text="read-only",
            results=(),
            retrieval_modes=("core", "fallback"),
        )

        result = chat_gui_bridge._serialize_memory_context(context)

        self.assertEqual(["core", "fallback"], result["retrieval_modes"])

    def test_send_message_uses_attachment_context_without_saving_body(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_text = "PRIVATE_ATTACHMENT_BODY"

            result = chat_gui_bridge.handle_send_message(
                {
                    "root": str(root),
                    "content": "添付を確認して",
                    "no_ai": True,
                    "attachments": [
                        {
                            "name": "notes.md",
                            "extension": "md",
                            "size_bytes": len(attachment_text.encode("utf-8")),
                            "text": attachment_text,
                        }
                    ],
                }
            )

            session_path = root / result["session"]["jsonl_file"]
            records = [json.loads(line) for line in session_path.read_text(encoding="utf-8").splitlines()]

            self.assertEqual("extracted", result["attachments"][0]["status"])
            self.assertIn("[Attachments]", records[0]["content"])
            self.assertIn("notes.md", records[0]["content"])
            self.assertNotIn(attachment_text, records[0]["content"])

    def test_send_message_returns_pdf_attachment_error_when_extraction_unavailable_or_invalid(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)

            result = chat_gui_bridge.handle_send_message(
                {
                    "root": str(root),
                    "content": "PDFを確認して",
                    "no_ai": True,
                    "attachments": [
                        {
                            "name": "sample.pdf",
                            "extension": "pdf",
                            "size_bytes": 12,
                            "data_base64": "bm90IGEgcGRm",
                        }
                    ],
                }
            )

            attachment = result["attachments"][0]
            self.assertEqual("error", attachment["status"])
            self.assertTrue(attachment["error"])

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_xlsx_attachment_extracts_formula_without_saving_body(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.append(["Item", "Amount"])
        sheet.append(["Coffee", "=SUM(200,250)"])
        hidden = workbook.create_sheet("Private")
        hidden["A1"] = "HIDDEN_VALUE"
        hidden.sheet_state = "hidden"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        raw = output.getvalue()
        encoded = base64.b64encode(raw).decode("ascii")

        report = chat_gui_bridge._normalize_attachment(
            {
                "name": "budget.xlsx",
                "size_bytes": len(raw),
                "data_base64": encoded,
            }
        )

        self.assertEqual("extracted", report["status"])
        self.assertIn("# Sheet: Summary", report["_text"])
        self.assertIn("Coffee\t=SUM(200,250)", report["_text"])
        self.assertNotIn("HIDDEN_VALUE", report["_text"])

        captured_prompt = {}
        original_generate = chat_gui_bridge.generate_assistant_reply_with_context

        def fake_generate(root, messages, **_kwargs):
            del root
            captured_prompt["latest"] = messages[-1].content
            return codex_conversation.AssistantReplyResult(reply="確認しました。", memory_context=None)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            chat_gui_bridge.generate_assistant_reply_with_context = fake_generate
            try:
                result = chat_gui_bridge.handle_send_message(
                    {
                        "root": str(root),
                        "content": "Excelを確認して",
                        "attachments": [
                            {
                                "name": "budget.xlsx",
                                "size_bytes": len(raw),
                                "data_base64": encoded,
                            }
                        ],
                    }
                )
            finally:
                chat_gui_bridge.generate_assistant_reply_with_context = original_generate
            session_path = root / result["session"]["jsonl_file"]
            saved = session_path.read_text(encoding="utf-8")

        self.assertIn("Coffee\t=SUM(200,250)", captured_prompt["latest"])
        self.assertIn("budget.xlsx", saved)
        self.assertNotIn("Coffee", saved)
        self.assertNotIn("=SUM(200,250)", saved)

    def test_xlsx_attachment_reports_extraction_limits(self):
        class FakeSheet:
            title = "Large"
            sheet_state = "visible"
            max_row = 1
            max_column = 1

            def iter_rows(self, **_kwargs):
                return iter([("x" * (chat_gui_bridge.MAX_ATTACHMENT_TEXT_CHARS + 100),)])

        class FakeWorkbook:
            worksheets = [FakeSheet()]

            def close(self):
                return None

        original_loader = chat_gui_bridge._load_xlsx_workbook
        chat_gui_bridge._load_xlsx_workbook = lambda _raw: FakeWorkbook()
        try:
            report = chat_gui_bridge._normalize_attachment(
                {
                    "name": "large.xlsx",
                    "size_bytes": 4,
                    "data_base64": base64.b64encode(b"xlsx").decode("ascii"),
                }
            )
        finally:
            chat_gui_bridge._load_xlsx_workbook = original_loader

        self.assertEqual("extracted", report["status"])
        self.assertTrue(report["truncated"])
        self.assertEqual(chat_gui_bridge.MAX_ATTACHMENT_TEXT_CHARS, report["extracted_chars"])

    def test_xlsx_attachment_returns_clear_dependency_error(self):
        original_loader = chat_gui_bridge._load_xlsx_workbook
        chat_gui_bridge._load_xlsx_workbook = lambda _raw: (_ for _ in ()).throw(
            RuntimeError("Excel抽出には openpyxl が必要です。")
        )
        try:
            report = chat_gui_bridge._normalize_attachment(
                {
                    "name": "missing.xlsx",
                    "size_bytes": 4,
                    "data_base64": base64.b64encode(b"xlsx").decode("ascii"),
                }
            )
        finally:
            chat_gui_bridge._load_xlsx_workbook = original_loader

        self.assertEqual("error", report["status"])
        self.assertIn("openpyxl", report["error"])

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_xlsx_attachment_returns_error_for_invalid_workbook(self):
        report = chat_gui_bridge._normalize_attachment(
            {
                "name": "broken.xlsx",
                "size_bytes": 12,
                "data_base64": base64.b64encode(b"not an xlsx").decode("ascii"),
            }
        )

        self.assertEqual("error", report["status"])
        self.assertIn("Excelファイルを開けませんでした", report["error"])

    def test_resume_session_returns_messages(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            now = datetime.now(timezone.utc) - timedelta(days=365)
            session = create_live_session(root=root, started_at=now)
            session.append_message("user", "resume me", now)
            session.append_message("assistant", "ok", now + timedelta(seconds=2))

            result = chat_gui_bridge.handle_resume_session(
                {
                    "root": str(root),
                    "session_ref": session.path.stem,
                }
            )

            self.assertEqual(result["session"]["session_id"], session.path.stem)
            self.assertEqual([message["role"] for message in result["messages"]], ["user", "assistant"])

    def test_list_resumable_sessions_includes_old_sessions(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            old = datetime.now(timezone.utc) - timedelta(days=365)
            session = create_live_session(root=root, started_at=old)
            session.append_message("user", "list me", old)

            result = chat_gui_bridge.handle_list_resumable({"root": str(root), "max_sessions": 50})

            self.assertEqual([session.path.stem], [item["session_id"] for item in result["sessions"]])

    def test_cleanup_expired_command_is_not_exposed(self):
        self.assertNotIn("cleanup-expired", chat_gui_bridge.COMMANDS)

    def test_finalize_job_runs_in_background_and_returns_status(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            prompt_dir = root / "prompts"
            prompt_dir.mkdir()
            (prompt_dir / "codex_phase2_prompt.md").write_text("Process {RAW_FILE}", encoding="utf-8")
            session = create_live_session(root=root, started_at=datetime(2026, 7, 3, 12, 0, tzinfo=timezone.utc))
            session.append_message("user", "finalize me", session.started_at)

            started = chat_gui_bridge.handle_start_finalize_job(
                {
                    "root": str(root),
                    "session_file": str(session.path),
                    "run_codex": False,
                }
            )
            job_id = started["job"]["job_id"]

            status = None
            for _ in range(50):
                status = chat_gui_bridge.handle_get_finalize_job({"root": str(root), "job_id": job_id})["job"]
                if status["status"] in {"succeeded", "failed", "cancelled"}:
                    break
                time.sleep(0.1)

            self.assertIsNotNone(status)
            self.assertEqual("succeeded", status["status"])
            self.assertTrue((root / status["result"]["raw_file"]).exists())
            for process in chat_gui_bridge.BACKGROUND_PROCESSES:
                process.wait(timeout=5)
            self.assertFalse(chat_gui_bridge._finalize_lock_path(root, session.path).exists())
            job_log = (root / status["log_path"]).read_text(encoding="utf-8")
            self.assertIn("worker.start", job_log)
            self.assertIn("worker.progress", job_log)
            self.assertIn("worker.succeeded", job_log)

    def test_finalize_job_reuses_active_job_for_same_session(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            session = create_live_session(root=root, started_at=datetime(2026, 7, 3, 12, 0, tzinfo=timezone.utc))
            session.append_message("user", "finalize once", session.started_at)
            fake_process = mock.Mock(pid=os.getpid())

            with mock.patch.object(chat_gui_bridge, "_spawn_finalize_worker", return_value=fake_process) as spawn:
                first = chat_gui_bridge.handle_start_finalize_job(
                    {"root": str(root), "session_file": str(session.path), "run_codex": False}
                )["job"]
                second = chat_gui_bridge.handle_start_finalize_job(
                    {"root": str(root), "session_file": str(session.path), "run_codex": False}
                )["job"]

            self.assertEqual(first["job_id"], second["job_id"])
            spawn.assert_called_once()
            self.assertEqual(1, len(list((root / "logs" / "chat_gui_jobs").glob("*.json"))))

    def test_finalize_job_lock_prevents_concurrent_duplicate_start(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            session = create_live_session(root=root, started_at=datetime(2026, 7, 3, 12, 0, tzinfo=timezone.utc))
            session.append_message("user", "concurrent finalize", session.started_at)
            payload = {"root": str(root), "session_file": str(session.path), "run_codex": False}
            fake_process = mock.Mock(pid=os.getpid())

            with mock.patch.object(chat_gui_bridge, "_spawn_finalize_worker", return_value=fake_process) as spawn:
                with ThreadPoolExecutor(max_workers=2) as executor:
                    jobs = list(executor.map(lambda _: chat_gui_bridge.handle_start_finalize_job(payload)["job"], range(2)))

            self.assertEqual(jobs[0]["job_id"], jobs[1]["job_id"])
            spawn.assert_called_once()

    def test_get_finalize_job_recovers_dead_worker_and_releases_lock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            session = create_live_session(root=root, started_at=datetime(2026, 7, 3, 12, 0, tzinfo=timezone.utc))
            session.append_message("user", "recover orphan", session.started_at)
            job_id = "orphan-job"
            chat_gui_bridge._write_job_status(
                root,
                job_id,
                {
                    "job_id": job_id,
                    "status": "running",
                    "stage": "memory",
                    "session_file": chat_gui_bridge._display_path(session.path, root),
                    "created_at": "2026-07-03T12:00:00+00:00",
                    "worker_pid": 999999,
                },
            )
            self.assertIsNone(chat_gui_bridge._claim_finalize_lock(root, session.path, job_id))

            with mock.patch.object(chat_gui_bridge, "_process_exists", return_value=False):
                status = chat_gui_bridge.handle_get_finalize_job({"root": str(root), "job_id": job_id})["job"]

            self.assertEqual("failed", status["status"])
            self.assertIn("terminal result", status["error"])
            self.assertFalse(chat_gui_bridge._finalize_lock_path(root, session.path).exists())
            job_log = chat_gui_bridge._job_log_path(root, job_id).read_text(encoding="utf-8")
            self.assertIn("worker.missing", job_log)

    def test_cancelable_runner_keeps_non_utf8_stderr_as_bytes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runner = chat_gui_bridge._cancelable_run_command(root, root / "cancel")
            completed = runner(
                [
                    sys.executable,
                    "-c",
                    "import sys; sys.stderr.buffer.write('認証エラー'.encode('cp932')); raise SystemExit(1)",
                ],
                cwd=root,
                input=b"",
                text=False,
            )

            self.assertEqual(1, completed.returncode)
            self.assertEqual("認証エラー".encode("cp932"), completed.stderr)

    def test_finalize_worker_uses_detached_windows_process_flags(self):
        flags = chat_gui_bridge._finalize_worker_creationflags()
        if os.name != "nt":
            self.assertEqual(0, flags)
            return

        self.assertNotEqual(0, flags & subprocess.DETACHED_PROCESS)
        self.assertNotEqual(0, flags & subprocess.CREATE_NEW_PROCESS_GROUP)
        self.assertNotEqual(0, flags & subprocess.CREATE_BREAKAWAY_FROM_JOB)

    def test_cancelable_subprocess_hides_windows_console(self):
        flags = chat_gui_bridge._cancelable_subprocess_creationflags()
        if os.name != "nt":
            self.assertEqual(0, flags)
            return

        self.assertNotEqual(0, flags & subprocess.CREATE_NEW_PROCESS_GROUP)
        self.assertNotEqual(0, flags & subprocess.CREATE_NO_WINDOW)

    def test_organize_sessions_job_processes_unorganized_sessions_oldest_first(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            prompt_dir = root / "prompts"
            prompt_dir.mkdir()
            (prompt_dir / "codex_phase2_prompt.md").write_text("Process {RAW_FILE}", encoding="utf-8")
            now = datetime.now(timezone.utc)
            older = create_live_session(root=root, started_at=now - timedelta(minutes=2))
            newer = create_live_session(root=root, started_at=now - timedelta(minutes=1))
            older.append_message("user", "older", older.started_at)
            newer.append_message("user", "newer", newer.started_at)

            started = chat_gui_bridge.handle_start_organize_sessions_job(
                {"root": str(root), "run_codex": False}
            )
            self.assertEqual(2, started["eligible_count"])
            job_id = started["job"]["job_id"]

            status = None
            for _ in range(50):
                status = chat_gui_bridge.handle_get_organize_sessions_job({"root": str(root), "job_id": job_id})["job"]
                if status["status"] in {"succeeded", "failed", "cancelled"}:
                    break
                time.sleep(0.1)

            self.assertIsNotNone(status)
            self.assertEqual("succeeded", status["status"])
            self.assertEqual([older.path.stem, newer.path.stem], status["result"]["completed_sessions"])
            for process in chat_gui_bridge.BACKGROUND_PROCESSES:
                process.wait(timeout=5)
            self.assertFalse(chat_gui_bridge._organize_sessions_lock_path(root).exists())
            job_log = (root / status["log_path"]).read_text(encoding="utf-8")
            self.assertIn("worker.start targets=2", job_log)
            self.assertIn("worker.succeeded", job_log)

    def test_organize_sessions_job_reuses_active_job(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            now = datetime.now(timezone.utc)
            session = create_live_session(root=root, started_at=now)
            session.append_message("user", "organize", now)
            fake_process = mock.Mock(pid=os.getpid())

            with mock.patch.object(chat_gui_bridge, "_spawn_organize_sessions_worker", return_value=fake_process) as spawn:
                first = chat_gui_bridge.handle_start_organize_sessions_job({"root": str(root)})["job"]
                second = chat_gui_bridge.handle_start_organize_sessions_job({"root": str(root)})["job"]

            self.assertEqual(first["job_id"], second["job_id"])
            spawn.assert_called_once()

    def test_active_organize_sessions_job_blocks_individual_finalize(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            now = datetime.now(timezone.utc)
            session = create_live_session(root=root, started_at=now)
            session.append_message("user", "blocked", now)
            job_id = "organize-active"
            chat_gui_bridge._write_job_status(
                root,
                job_id,
                {
                    "job_id": job_id,
                    "name": chat_gui_bridge.ORGANIZE_SESSIONS_JOB_NAME,
                    "status": "running",
                    "created_at": now.isoformat(),
                    "worker_pid": os.getpid(),
                },
            )

            with self.assertRaisesRegex(RuntimeError, "データ整理が進行中"):
                chat_gui_bridge.handle_start_finalize_job({"root": str(root), "session_file": str(session.path)})

    def test_get_organize_sessions_job_recovers_dead_worker_and_releases_lock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            job_id = "organize-orphan"
            chat_gui_bridge._write_job_status(
                root,
                job_id,
                {
                    "job_id": job_id,
                    "name": chat_gui_bridge.ORGANIZE_SESSIONS_JOB_NAME,
                    "status": "running",
                    "created_at": "2026-07-03T12:00:00+00:00",
                    "worker_pid": 999999,
                },
            )
            self.assertIsNone(chat_gui_bridge._claim_organize_sessions_lock(root, job_id))

            with mock.patch.object(chat_gui_bridge, "_process_exists", return_value=False):
                status = chat_gui_bridge.handle_get_organize_sessions_job({"root": str(root), "job_id": job_id})["job"]

            self.assertEqual("failed", status["status"])
            self.assertIn("terminal result", status["error"])
            self.assertFalse(chat_gui_bridge._organize_sessions_lock_path(root).exists())


if __name__ == "__main__":
    unittest.main()
