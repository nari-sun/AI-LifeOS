import argparse
import base64
import binascii
import hashlib
import json
import os
import re
import subprocess
import sys
import time
import traceback
from datetime import date, datetime, time as datetime_time, timezone
from pathlib import Path
from typing import Any, Callable
from uuid import uuid4

from build_answer_context import AnswerContext, MemoryContextReference
from codex_conversation import (
    AppServerStreamingUnavailable,
    generate_assistant_reply_streaming_with_context,
    generate_assistant_reply_with_context,
)
from finalize_live_chat import finalize_live_chat
from import_chatgpt_export import (
    CONVERSATION_FILE_PATTERN,
    ExportConversation,
    IMPORT_STATE_CONFLICT,
    IMPORT_STATE_DUPLICATE,
    IMPORT_STATE_NEW,
    IMPORT_STATE_UPDATED,
    classify_import_states,
    import_conversations,
    load_export,
)
from live_session import ROOT, LiveMessage, LiveSession, create_live_message, create_live_session
from local_data_report import build_local_data_report
from memory_index import inspect_index_health, rebuild_index
from kokoro_tts import (
    DEFAULT_VOICE,
    KokoroSynthesisCancelled,
    KokoroTtsError,
    SUPPORTED_VOICES,
    cleanup_temp_audio,
    synthesize_to_wav,
    synthesize_to_wav_chunks,
    temporary_audio_dir,
)
from personalization_settings import (
    build_memory_summary,
    load_personalization_settings,
    load_session_personalization,
    personalization_settings_path,
    serialize_personalization,
    update_personalization_settings,
    update_session_personalization,
    validate_project_scope,
)
from session_store import get_session_organization, list_resumable_sessions, load_resume_session, save_session

GUI_LOG_ENV = "AI_LIFEOS_GUI_LOG"
REQUEST_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]+$")
MAX_ATTACHMENTS = 3
MAX_ATTACHMENT_BYTES = 1024 * 1024
MAX_ATTACHMENT_TEXT_CHARS = 12_000
MAX_XLSX_SHEETS = 20
MAX_XLSX_ROWS_PER_SHEET = 200
MAX_XLSX_COLUMNS_PER_SHEET = 50
ALLOWED_ATTACHMENT_EXTENSIONS = {".txt", ".md", ".pdf", ".xlsx"}
TEXT_ATTACHMENT_EXTENSIONS = {".txt", ".md"}
JOB_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{1,128}$")
LOCAL_DATA_FOLDERS = {"conversations", "journal", "memory", "inbox", "tasks", "imports", "logs"}
BACKGROUND_PROCESSES: list[subprocess.Popen[str]] = []
FINALIZE_ACTIVE_STATUSES = {"queued", "running"}
FINALIZE_TERMINAL_STATUSES = {"succeeded", "failed", "cancelled"}
FINALIZE_WORKER_START_GRACE_SECONDS = 5.0
ORGANIZE_SESSIONS_JOB_NAME = "organize-sessions"
READ_ALOUD_REQUEST_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{1,128}$")


class AssistantGenerationCancelled(RuntimeError):
    pass


def handle_start_session(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    session = create_live_session(root=root)
    settings = load_personalization_settings(root)
    update_session_personalization(
        root,
        session.path,
        temporary=(payload["temporary"] if "temporary" in payload else False),
        memory_enabled=settings.memory_enabled,
        past_chat_search_enabled=settings.past_chat_search_enabled,
        project_scope=(payload["project_scope"] if "project_scope" in payload else settings.project_scope),
    )
    _gui_log(root, f"start_session.created session={session.path.stem}")
    return {
        "session": _serialize_session_file(session.path, root),
        "messages": [],
    }


def handle_send_message(
    payload: dict[str, Any],
    on_delta: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    root = _payload_root(payload)
    content = str(payload.get("content", "")).strip()
    if not content:
        raise ValueError("送信するメッセージが空です。")
    full_archive_review = payload.get("full_archive_review", False)
    if not isinstance(full_archive_review, bool):
        raise ValueError("full_archive_review は true または false で指定してください。")

    session_file = _resolve_or_create_session(root=root, value=payload.get("session_file"))
    request_id = _optional_request_id(payload.get("request_id"))
    cancel_file = _cancel_file_path(root, request_id) if request_id else None
    if cancel_file:
        _clear_cancel_file(cancel_file)
    _gui_log(
        root,
        f"send_message.start session={session_file.stem} request_id={request_id or '-'} no_ai={bool(payload.get('no_ai', False))}",
    )

    session = LiveSession(path=session_file, started_at=_session_started_at(session_file))
    messages = _read_live_messages(session_file)
    attachments = _normalize_attachments(payload.get("attachments"))
    attachment_metadata = _format_attachment_metadata_for_saved_message(attachments)
    saved_user_content = content + attachment_metadata
    prompt_user_content = _format_user_content_for_generation(content, attachments)

    user_message = create_live_message("user", saved_user_content)
    session.append_message(user_message.role, user_message.content, user_message.timestamp)
    messages.append(user_message)
    _save_session_metadata(root=root, session_file=session_file, status="saved")

    # Read the effective snapshot only after the first user record is durable.
    # A temporary-mode update that won the race before the append is therefore
    # honored by this reply; one that loses the race is rejected by the settings
    # layer because the session is now locked.
    personalization = load_session_personalization(root, session_file)
    retrieval_enabled = personalization.memory_enabled or personalization.past_chat_search_enabled

    assistant_message = None
    memory_context = None
    memory_candidates: tuple[MemoryContextReference, ...] = ()
    memory_opened: tuple[MemoryContextReference, ...] = ()
    error = None
    cancelled = False
    if not bool(payload.get("no_ai", False)):
        try:
            run_command = _cancelable_run_command(root=root, cancel_file=cancel_file) if cancel_file else subprocess.run
            generation_messages = _messages_for_generation(messages, user_message, prompt_user_content)
            generation_options = {
                "root": root,
                "messages": generation_messages,
                "codex_command": str(payload.get("codex_command") or "codex.cmd"),
                "sandbox": str(payload.get("codex_sandbox") or "read-only"),
                "approval": str(payload.get("codex_approval") or "never"),
                "max_context_messages": int(payload.get("max_context_messages") or 20),
                "include_memory_context": retrieval_enabled,
                "enable_memory_mcp": personalization.past_chat_search_enabled,
                "include_core_memory": personalization.memory_enabled,
                "include_past_chats": personalization.past_chat_search_enabled,
                "force_full_archive_review": full_archive_review,
                "project_scope": personalization.project_scope,
                "exclude_live_session": session_file,
            }
            if on_delta is None:
                reply_result = generate_assistant_reply_with_context(
                    **generation_options,
                    run_command=run_command,
                )
            else:
                try:
                    reply_result = generate_assistant_reply_streaming_with_context(
                        **generation_options,
                        on_delta=on_delta,
                        is_cancelled=(lambda: bool(cancel_file and cancel_file.exists())),
                    )
                except AppServerStreamingUnavailable as exc:
                    _gui_log(
                        root,
                        "send_message.streaming_fallback "
                        f"session={session_file.stem} request_id={request_id or '-'} message={_safe_log_text(str(exc))}",
                    )
                    reply_result = generate_assistant_reply_with_context(
                        **generation_options,
                        run_command=run_command,
                    )
            reply = reply_result.reply
            memory_context = reply_result.memory_context
            candidate_values = getattr(reply_result, "memory_candidates", ())
            if isinstance(candidate_values, (list, tuple)):
                memory_candidates = tuple(candidate_values)
            opened_values = getattr(reply_result, "memory_opened", ())
            if isinstance(opened_values, (list, tuple)):
                memory_opened = tuple(opened_values)
            if memory_context is not None:
                raw_chunk_count = sum(1 for result in memory_context.results if result.document_type == "raw_chunk")
                _gui_log(
                    root,
                    "send_message.memory_context "
                    f"session={session_file.stem} used={memory_context.used_memory} "
                    f"score={memory_context.score}/{memory_context.threshold} "
                    f"references={len(memory_context.references)} results={len(memory_context.results)} "
                    f"raw_chunks={raw_chunk_count}",
                )
            if cancel_file and cancel_file.exists():
                raise AssistantGenerationCancelled("返答生成を停止しました。")
            saved_assistant = create_live_message("assistant", reply)
            session.append_message(saved_assistant.role, saved_assistant.content, saved_assistant.timestamp)
            messages.append(saved_assistant)
            assistant_message = saved_assistant
            _save_session_metadata(root=root, session_file=session_file, status="saved")
        except (AssistantGenerationCancelled, InterruptedError) as exc:
            cancelled = True
            _gui_log(
                root,
                f"send_message.cancelled session={session_file.stem} request_id={request_id or '-'} message={_safe_log_text(str(exc))}",
            )
        except Exception as exc:  # Keep the already-saved user message visible to the GUI.
            error = f"{type(exc).__name__}: {exc}"
            _gui_log(root, f"send_message.assistant_error session={session_file.stem} {_format_exception(exc)}")
        finally:
            if cancel_file:
                _clear_cancel_file(cancel_file)

    _gui_log(
        root,
        f"send_message.done session={session_file.stem} messages={len(messages)} assistant_saved={assistant_message is not None} cancelled={cancelled}",
    )
    return {
        "session": _serialize_session_file(session_file, root),
        "messages": [_serialize_message(message) for message in messages],
        "assistant": _serialize_message(assistant_message) if assistant_message else None,
        "memory_context": _serialize_memory_context(memory_context),
        "memory_candidates": [_serialize_memory_reference(reference) for reference in memory_candidates],
        "memory_opened": [_serialize_memory_reference(reference) for reference in memory_opened],
        "attachments": [_serialize_attachment_report(attachment) for attachment in attachments],
        "error": error,
        "cancelled": cancelled,
    }


def handle_read_aloud(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload).resolve()
    request_id = _read_aloud_request_id(payload.get("request_id"))
    voice = str(payload.get("voice") or DEFAULT_VOICE).strip()
    if voice not in SUPPORTED_VOICES:
        raise ValueError("選択した読み上げ voice は利用できません。")

    cancel_file = _read_aloud_cancel_file(root, request_id)
    cancel_file.parent.mkdir(parents=True, exist_ok=True)
    if cancel_file.exists():
        cancel_file.unlink(missing_ok=True)
        raise KokoroSynthesisCancelled("読み上げを停止しました。")
    _tts_log(root, f"read_aloud.start request_id={request_id} voice={voice} chars={len(str(payload.get('text') or ''))}")
    try:
        audio_path = synthesize_to_wav(
            root=root,
            text=str(payload.get("text") or ""),
            voice=voice,
            request_id=request_id,
            is_cancelled=cancel_file.exists,
        )
    except KokoroSynthesisCancelled:
        _tts_log(root, f"read_aloud.cancelled request_id={request_id}")
        raise
    except KokoroTtsError as exc:
        _tts_log(root, f"read_aloud.failed request_id={request_id} type={type(exc).__name__}")
        raise
    finally:
        cancel_file.unlink(missing_ok=True)

    _tts_log(root, f"read_aloud.done request_id={request_id} voice={voice}")
    return {
        "request_id": request_id,
        "voice": voice,
        "audio_file": _display_path(audio_path, root),
        "audio_path": str(audio_path.resolve()),
    }


def handle_read_aloud_stream(payload: dict[str, Any], on_audio: Callable[[dict[str, Any]], None]) -> dict[str, Any]:
    root = _payload_root(payload).resolve()
    request_id = _read_aloud_request_id(payload.get("request_id"))
    voice = str(payload.get("voice") or DEFAULT_VOICE).strip()
    if voice not in SUPPORTED_VOICES:
        raise ValueError("選択した読み上げ voice は利用できません。")

    cancel_file = _read_aloud_cancel_file(root, request_id)
    cancel_file.parent.mkdir(parents=True, exist_ok=True)
    if cancel_file.exists():
        cancel_file.unlink(missing_ok=True)
        raise KokoroSynthesisCancelled("読み上げを停止しました。")
    _tts_log(root, f"read_aloud_stream.start request_id={request_id} voice={voice} chars={len(str(payload.get('text') or ''))}")

    def publish_audio(audio_path: Path, index: int) -> None:
        on_audio(
            {
                "request_id": request_id,
                "voice": voice,
                "index": index,
                "audio_file": _display_path(audio_path, root),
                "audio_path": str(audio_path.resolve()),
            }
        )
        _tts_log(root, f"read_aloud_stream.chunk request_id={request_id} index={index}")

    try:
        audio_paths = synthesize_to_wav_chunks(
            root=root,
            text=str(payload.get("text") or ""),
            voice=voice,
            request_id=request_id,
            is_cancelled=cancel_file.exists,
            on_chunk=publish_audio,
        )
    except KokoroSynthesisCancelled:
        _tts_log(root, f"read_aloud_stream.cancelled request_id={request_id}")
        raise
    except KokoroTtsError as exc:
        _tts_log(root, f"read_aloud_stream.failed request_id={request_id} type={type(exc).__name__}")
        raise
    finally:
        cancel_file.unlink(missing_ok=True)

    _tts_log(root, f"read_aloud_stream.done request_id={request_id} chunks={len(audio_paths)}")
    return {"request_id": request_id, "voice": voice, "chunk_count": len(audio_paths)}


def handle_cancel_read_aloud(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload).resolve()
    request_id = _read_aloud_request_id(payload.get("request_id"))
    cancel_file = _read_aloud_cancel_file(root, request_id)
    cancel_file.parent.mkdir(parents=True, exist_ok=True)
    cancel_file.write_text("cancel\n", encoding="utf-8")
    _tts_log(root, f"read_aloud.cancel_requested request_id={request_id}")
    return {"request_id": request_id, "cancelled": True}


def handle_discard_read_aloud_audio(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload).resolve()
    audio_path = _read_aloud_audio_path(root, payload.get("audio_path"))
    removed = False
    try:
        audio_path.unlink(missing_ok=True)
        removed = True
    except OSError:
        removed = False
    cleanup_temp_audio(_read_aloud_runtime_dir())
    _tts_log(root, f"read_aloud.discard file={audio_path.name} removed={removed}")
    return {"audio_file": _display_path(audio_path, root), "removed": removed}


def handle_cancel_message(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    request_id = _required_request_id(payload.get("request_id"))
    cancel_file = _cancel_file_path(root, request_id)
    cancel_file.parent.mkdir(parents=True, exist_ok=True)
    cancel_file.write_text(datetime.now().astimezone().isoformat(timespec="seconds"), encoding="utf-8")
    _gui_log(root, f"cancel_message.requested request_id={request_id}")
    return {
        "request_id": request_id,
        "cancelled": True,
    }


def handle_save_session(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    session_file = _resolve_existing_session(root=root, value=payload.get("session_file"))
    _gui_log(root, f"save_session.start session={session_file.stem}")
    saved = save_session(
        root=root,
        session_file=session_file,
        title=_optional_text(payload.get("title")),
        status="saved",
    )
    _gui_log(root, f"save_session.done session={saved.session_id} messages={saved.message_count}")
    return {"saved": _serialize_saved_session(saved, root)}


def handle_list_resumable(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    max_sessions = int(payload.get("max_sessions") or 50)
    sessions = list_resumable_sessions(root=root, limit=max_sessions)
    _gui_log(root, f"list_resumable.done limit={max_sessions} count={len(sessions)}")
    return {"sessions": [_serialize_resume_session(session, root) for session in sessions]}


def handle_resume_session(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    session_ref = str(payload.get("session_ref") or "latest")
    _gui_log(root, f"resume_session.start ref={_safe_log_text(session_ref)}")
    summary, records = load_resume_session(root=root, session_ref=session_ref)
    messages = [_message_from_record(record) for record in records]
    _gui_log(root, f"resume_session.done session={summary.session_id} messages={len(messages)}")
    return {
        "session": _serialize_session_file(summary.jsonl_file, root),
        "summary": _serialize_resume_session(summary, root),
        "messages": [_serialize_message(message) for message in messages],
    }


def handle_get_personalization(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    settings = load_personalization_settings(root)
    session_value = payload.get("session_file")
    session_file = _resolve_session_path(root=root, value=session_value, must_exist=False) if session_value else None
    session = load_session_personalization(root, session_file) if session_file else None
    _gui_log(root, f"personalization.get session={Path(str(session_value)).stem if session_value else '-'}")
    return {
        "settings": serialize_personalization(settings),
        "session": serialize_personalization(session) if session else None,
        "session_state": _serialize_session_file(session_file, root) if session_file else None,
        "settings_file": _display_path(personalization_settings_path(root), root),
    }


def handle_update_personalization(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    settings_payload = payload.get("settings")
    session_payload = payload.get("session")
    if settings_payload is None and session_payload is None:
        raise ValueError("更新するパーソナライズ設定を指定してください。")

    settings_fields = None
    if settings_payload is not None:
        settings_fields = _personalization_update_fields(
            settings_payload,
            allowed={"memory_enabled", "past_chat_search_enabled", "project_scope"},
            label="settings",
        )

    session = None
    session_file = None
    session_fields = None
    if session_payload is not None:
        session_file = _resolve_session_path(root=root, value=payload.get("session_file"), must_exist=False)
        session_fields = _personalization_update_fields(
            session_payload,
            allowed={"temporary", "memory_enabled", "past_chat_search_enabled", "project_scope"},
            label="session",
        )
        if session_fields.get("temporary") is True and (
            _find_active_finalize_job(root=root, session_file=session_file) is not None
            or _find_active_organize_sessions_job(root) is not None
        ):
            raise RuntimeError("整理ジョブの実行中は一時チャットへ変更できません。整理完了後に新規セッションで指定してください。")
        # The settings layer owns the immutable retention boundary check for
        # both directions (normal -> temporary and temporary -> normal).

    settings = (
        update_personalization_settings(root, **settings_fields)
        if settings_fields is not None
        else load_personalization_settings(root)
    )
    if session_fields is not None and session_file is not None:
        session = update_session_personalization(root, session_file, **session_fields)

    _gui_log(
        root,
        "personalization.update "
        f"settings={settings_payload is not None} session={session_payload is not None} "
        f"temporary={session.temporary if session else '-'}",
    )
    return {
        "settings": serialize_personalization(settings),
        "session": serialize_personalization(session) if session else None,
        "session_state": _serialize_session_file(session_file, root) if session is not None else None,
        "settings_file": _display_path(personalization_settings_path(root), root),
    }


def handle_get_memory_summary(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    summary = build_memory_summary(root)
    _gui_log(
        root,
        "memory_summary.get "
        f"sections={len(summary['sections'])} structured_items={len(summary['structured_items'])}",
    )
    return {"summary": summary}


def handle_finalize_session(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    if _find_active_organize_sessions_job(root) is not None:
        raise RuntimeError("データ整理が進行中です。完了または停止してから個別整理を実行してください。")
    session_file = _resolve_existing_session(root=root, value=payload.get("session_file"))
    _ensure_session_can_be_organized(root, session_file)
    run_codex = bool(payload.get("run_codex", True))
    _gui_log(root, f"finalize_session.start session={session_file.stem} run_codex={run_codex}")
    result = finalize_live_chat(
        root=root,
        session_file=session_file,
        run_codex=run_codex,
        commit=False,
        force=True,
        codex_command=str(payload.get("codex_command") or "codex.cmd"),
        codex_sandbox="workspace-write",
        codex_approval=str(payload.get("codex_approval") or "never"),
    )
    _gui_log(root, f"finalize_session.done session={session_file.stem} raw={_display_path(result.raw_file, root)}")
    return {
        "session": _serialize_session_file(result.jsonl_file, root),
        "jsonl_file": _display_path(result.jsonl_file, root),
        "raw_file": _display_path(result.raw_file, root),
        "task_file": _display_path(result.task_file, root),
        "imported_at": result.imported_at.isoformat(timespec="seconds"),
        "codex_updated": bool(result.codex),
        "git_committed": bool(result.git and result.git.committed),
        "organization": result.organization,
    }


def handle_local_data_report(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    _gui_log(root, "local_data_report.start")
    report = build_local_data_report(root=root)
    _gui_log(root, "local_data_report.done")
    return {"report": report}


def handle_open_local_data_folder(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    folder = str(payload.get("folder") or "").strip()
    if folder not in LOCAL_DATA_FOLDERS:
        raise ValueError("開けるフォルダはローカルデータ管理対象に限定されています。")

    path = (root / folder).resolve()
    try:
        path.relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError("フォルダはAI-LifeOSルート内を指定してください。") from exc
    if not path.exists() or not path.is_dir():
        raise FileNotFoundError(f"フォルダが見つかりません: {folder}")

    _open_folder(path)
    _gui_log(root, f"open_local_data_folder.done folder={folder}")
    return {"folder": folder, "path": _display_path(path, root)}


def handle_preview_chatgpt_import(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    source_path = _resolve_chatgpt_export_source(payload.get("source"))
    source = _load_chatgpt_export_source(source_path)
    import_states = classify_import_states(source.conversations, root / "conversations")
    conversations = [
        _serialize_chatgpt_import_conversation(
            item,
            import_state=import_states.get(item.source_id, IMPORT_STATE_CONFLICT),
        )
        for item in source.conversations
    ]
    state_counts = {
        state: sum(1 for item in conversations if item["import_state"] == state)
        for state in (
            IMPORT_STATE_NEW,
            IMPORT_STATE_UPDATED,
            IMPORT_STATE_DUPLICATE,
            IMPORT_STATE_CONFLICT,
        )
    }
    _gui_log(
        root,
        f"chatgpt_import.preview conversations={len(conversations)} "
        f"new={state_counts[IMPORT_STATE_NEW]} updated={state_counts[IMPORT_STATE_UPDATED]} "
        f"duplicates={state_counts[IMPORT_STATE_DUPLICATE]} conflicts={state_counts[IMPORT_STATE_CONFLICT]}",
    )
    return {
        "source": source.display_name,
        "total_count": len(conversations),
        "new_count": state_counts[IMPORT_STATE_NEW],
        "updated_count": state_counts[IMPORT_STATE_UPDATED],
        "duplicate_count": state_counts[IMPORT_STATE_DUPLICATE],
        "conflict_count": state_counts[IMPORT_STATE_CONFLICT],
        "conversations": conversations,
    }


def handle_apply_chatgpt_import(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("confirmed") is not True:
        raise ValueError("取り込みにはGUIでの最終確認が必要です。")

    root = _payload_root(payload)
    source_path = _resolve_chatgpt_export_source(payload.get("source"))
    source = _load_chatgpt_export_source(source_path)
    selected_ids = _selected_chatgpt_import_ids(payload.get("selected_ids"))
    known_ids = {item.source_id for item in source.conversations}
    unknown_ids = selected_ids - known_ids
    if unknown_ids:
        raise ValueError("選択した会話がエクスポート内に見つかりません。内容を再確認してください。")

    selected = [item for item in source.conversations if item.source_id in selected_ids]
    import_states = classify_import_states(selected, root / "conversations")
    if any(import_states.get(item.source_id) == IMPORT_STATE_CONFLICT for item in selected):
        raise ValueError("競合している会話は自動更新できません。既存のimport metadataを確認してください。")
    results = import_conversations(
        selected,
        conversations_dir=root / "conversations",
        source_display_name=source.display_name,
    )
    created = [item for item in results if not item.duplicate and not item.updated]
    updated = [item for item in results if item.updated]
    written = [item for item in results if not item.duplicate]
    duplicate_count = sum(1 for item in results if item.duplicate)
    index_updated, index_status, index_error = _rebuild_chatgpt_import_index(root)
    _gui_log(
        root,
        f"chatgpt_import.apply selected={len(selected)} imported={len(created)} updated={len(updated)} "
        f"duplicates={duplicate_count} index_updated={index_updated} index_status={index_status}",
    )
    return {
        "source": source.display_name,
        "selected_count": len(selected),
        "imported_count": len(created),
        "updated_count": len(updated),
        "duplicate_count": duplicate_count,
        "index_updated": index_updated,
        "index_status": index_status,
        "index_error": index_error,
        "imported": [
            {
                "source_id": item.conversation.source_id,
                "title": item.conversation.title,
                "raw_file": _display_path(item.raw_file, root) if item.raw_file else None,
                "updated": item.updated,
            }
            for item in written
        ],
    }


def handle_start_finalize_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    if _find_active_organize_sessions_job(root) is not None:
        raise RuntimeError("データ整理が進行中です。完了または停止してから個別整理を実行してください。")
    session_file = _resolve_existing_session(root=root, value=payload.get("session_file"))
    _ensure_session_can_be_organized(root, session_file)
    existing = _find_active_finalize_job(root=root, session_file=session_file)
    if existing is not None:
        _gui_log(root, f"finalize_job.reused job_id={existing['job_id']} session={session_file.stem}")
        return {"job": existing}

    job_id = _new_job_id()
    locked_job = _claim_finalize_lock(root=root, session_file=session_file, job_id=job_id)
    if locked_job is not None:
        _gui_log(root, f"finalize_job.reused job_id={locked_job['job_id']} session={session_file.stem}")
        return {"job": locked_job}

    log_path = _job_log_path(root, job_id)
    cancel_file = _job_cancel_file(root, job_id)
    _write_job_status(
        root,
        job_id,
        {
            "job_id": job_id,
            "name": "finalize-session",
            "status": "queued",
            "stage": "queued",
            "message": "整理ジョブを開始待ちです。",
            "error": None,
            "percent": 0,
            "session_file": _display_path(session_file, root),
            "log_path": _display_path(log_path, root),
            "cancel_file": _display_path(cancel_file, root),
            "created_at": _now_iso(),
            "started_at": None,
            "finished_at": None,
            "launcher_pid": os.getpid(),
            "worker_pid": None,
            "result": None,
        },
    )

    worker_payload = {
        "root": str(root),
        "job_id": job_id,
        "session_file": _display_path(session_file, root),
        "run_codex": bool(payload.get("run_codex", True)),
        "codex_command": str(payload.get("codex_command") or "codex.cmd"),
        "codex_approval": str(payload.get("codex_approval") or "never"),
    }
    try:
        process = _spawn_finalize_worker(root=root, payload=worker_payload, log_path=log_path)
        _write_job_status(root, job_id, {"worker_pid": process.pid})
    except Exception as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "failed",
                "stage": "failed",
                "message": "整理ワーカーを起動できませんでした。",
                "error": f"{type(exc).__name__}: {exc}",
                "finished_at": _now_iso(),
            },
        )
        _release_finalize_lock(root=root, session_file=session_file, job_id=job_id)
        raise
    _gui_log(root, f"finalize_job.started job_id={job_id} session={session_file.stem}")
    return {"job": _read_job_status(root, job_id)}


def handle_get_finalize_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    return {"job": _refresh_finalize_job_status(root, job_id)}


def handle_cancel_finalize_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    status = _refresh_finalize_job_status(root, job_id)
    if status.get("status") not in FINALIZE_TERMINAL_STATUSES:
        cancel_file = _job_cancel_file(root, job_id)
        cancel_file.parent.mkdir(parents=True, exist_ok=True)
        cancel_file.write_text(_now_iso(), encoding="utf-8")
        status = _write_job_status(
            root,
            job_id,
            {
                "status": status.get("status", "running"),
                "message": "キャンセル要求を送信しました。",
                "cancel_requested_at": _now_iso(),
            },
        )
        _gui_log(root, f"finalize_job.cancel_requested job_id={job_id}")
    return {"job": status}


def handle_run_finalize_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    _run_finalize_job(root=root, payload=payload, job_id=job_id)
    return {"job": _read_job_status(root, job_id)}


def handle_start_organize_sessions_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    existing = _find_active_organize_sessions_job(root)
    if existing is not None:
        _gui_log(root, f"organize_sessions.reused job_id={existing['job_id']}")
        return {"job": existing, "eligible_count": int(existing.get("total_sessions") or 0)}

    active_finalize = _find_any_active_finalize_job(root)
    if active_finalize is not None:
        raise RuntimeError("個別の整理ジョブが進行中です。完了または停止してからデータ整理を実行してください。")

    targets = _organize_session_targets(root=root)
    if not targets:
        _gui_log(root, "organize_sessions.none")
        return {"job": None, "eligible_count": 0}

    job_id = _new_job_id()
    locked_job = _claim_organize_sessions_lock(root=root, job_id=job_id)
    if locked_job is not None:
        _gui_log(root, f"organize_sessions.reused job_id={locked_job['job_id']}")
        return {"job": locked_job, "eligible_count": int(locked_job.get("total_sessions") or 0)}
    log_path = _job_log_path(root, job_id)
    cancel_file = _job_cancel_file(root, job_id)
    _write_job_status(
        root,
        job_id,
        {
            "job_id": job_id,
            "name": ORGANIZE_SESSIONS_JOB_NAME,
            "status": "queued",
            "stage": "queued",
            "message": "データ整理ジョブを開始待ちです。",
            "error": None,
            "percent": 0,
            "total_sessions": len(targets),
            "completed_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "failed_sessions": [],
            "current_session": None,
            "log_path": _display_path(log_path, root),
            "cancel_file": _display_path(cancel_file, root),
            "created_at": _now_iso(),
            "started_at": None,
            "finished_at": None,
            "launcher_pid": os.getpid(),
            "worker_pid": None,
            "result": None,
        },
    )

    worker_payload = {
        "root": str(root),
        "job_id": job_id,
        "session_files": [_display_path(path, root) for path in targets],
        "run_codex": bool(payload.get("run_codex", True)),
        "codex_command": str(payload.get("codex_command") or "codex.cmd"),
        "codex_approval": str(payload.get("codex_approval") or "never"),
    }
    try:
        process = _spawn_organize_sessions_worker(root=root, payload=worker_payload, log_path=log_path)
        _write_job_status(root, job_id, {"worker_pid": process.pid})
    except Exception as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "failed",
                "stage": "failed",
                "message": "データ整理ワーカーを起動できませんでした。",
                "error": f"{type(exc).__name__}: {exc}",
                "finished_at": _now_iso(),
            },
        )
        _release_organize_sessions_lock(root=root, job_id=job_id)
        raise
    _gui_log(root, f"organize_sessions.started job_id={job_id} targets={len(targets)}")
    return {"job": _read_job_status(root, job_id), "eligible_count": len(targets)}


def handle_get_organize_sessions_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    return {"job": _refresh_organize_sessions_job_status(root, job_id)}


def handle_cancel_organize_sessions_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    status = _refresh_organize_sessions_job_status(root, job_id)
    if status.get("status") not in FINALIZE_TERMINAL_STATUSES:
        cancel_file = _job_cancel_file(root, job_id)
        cancel_file.parent.mkdir(parents=True, exist_ok=True)
        cancel_file.write_text(_now_iso(), encoding="utf-8")
        status = _write_job_status(
            root,
            job_id,
            {
                "message": "データ整理の停止要求を送信しました。現在のセッション処理後に停止します。",
                "cancel_requested_at": _now_iso(),
            },
        )
        _gui_log(root, f"organize_sessions.cancel_requested job_id={job_id}")
    return {"job": status}


def handle_run_organize_sessions_job(payload: dict[str, Any]) -> dict[str, Any]:
    root = _payload_root(payload)
    job_id = _required_job_id(payload.get("job_id"))
    _run_organize_sessions_job(root=root, payload=payload, job_id=job_id)
    return {"job": _read_job_status(root, job_id)}


COMMANDS = {
    "start-session": handle_start_session,
    "read-aloud": handle_read_aloud,
    "read-aloud-stream": handle_read_aloud_stream,
    "cancel-read-aloud": handle_cancel_read_aloud,
    "discard-read-aloud-audio": handle_discard_read_aloud_audio,
    "send-message": handle_send_message,
    "send-message-stream": handle_send_message,
    "cancel-message": handle_cancel_message,
    "save-session": handle_save_session,
    "list-resumable": handle_list_resumable,
    "resume-session": handle_resume_session,
    "get-personalization": handle_get_personalization,
    "update-personalization": handle_update_personalization,
    "get-memory-summary": handle_get_memory_summary,
    "finalize-session": handle_finalize_session,
    "local-data-report": handle_local_data_report,
    "open-local-data-folder": handle_open_local_data_folder,
    "preview-chatgpt-import": handle_preview_chatgpt_import,
    "apply-chatgpt-import": handle_apply_chatgpt_import,
    "start-finalize-job": handle_start_finalize_job,
    "get-finalize-job": handle_get_finalize_job,
    "cancel-finalize-job": handle_cancel_finalize_job,
    "run-finalize-job": handle_run_finalize_job,
    "start-organize-sessions-job": handle_start_organize_sessions_job,
    "get-organize-sessions-job": handle_get_organize_sessions_job,
    "cancel-organize-sessions-job": handle_cancel_organize_sessions_job,
    "run-organize-sessions-job": handle_run_organize_sessions_job,
}


def main() -> int:
    _configure_stdio()
    parser = argparse.ArgumentParser(description="Bridge between the Tauri GUI and AI-LifeOS Python workflows.")
    parser.add_argument("command", choices=sorted(COMMANDS))
    args = parser.parse_args()

    payload: dict[str, Any] = {}
    root = ROOT
    try:
        payload = _read_payload()
        root = _payload_root(payload)
        _gui_log(root, f"command.start name={args.command} payload={_payload_log_summary(payload)}")
        if args.command == "send-message-stream":
            result = handle_send_message(payload, on_delta=_write_stream_delta)
            _write_json({"type": "result", "data": {"ok": True, **result}})
            return 0
        if args.command == "read-aloud-stream":
            result = handle_read_aloud_stream(payload, on_audio=_write_stream_audio)
            _write_json({"type": "result", "data": {"ok": True, **result}})
            return 0
        result = COMMANDS[args.command](payload)
        _gui_log(root, f"command.success name={args.command} result={_result_log_summary(result)}")
        _write_json({"ok": True, **result})
        return 0
    except Exception as exc:
        _gui_log(root, f"command.error name={args.command} {_format_exception(exc)}")
        _gui_log(root, "command.traceback\n" + traceback.format_exc())
        error_result = {"ok": False, "error": str(exc), "type": type(exc).__name__}
        if args.command in {"send-message-stream", "read-aloud-stream"}:
            _write_json({"type": "result", "data": error_result})
        else:
            _write_json(error_result)
        return 1


def _configure_stdio() -> None:
    for stream in (sys.stdin, sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")


def _read_payload() -> dict[str, Any]:
    raw = sys.stdin.read().lstrip("\ufeff")
    if not raw.strip():
        return {}

    value = json.loads(raw)
    if not isinstance(value, dict):
        raise ValueError("JSON payload must be an object.")
    return value


def _write_json(value: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(value, ensure_ascii=False, separators=(",", ":")))
    sys.stdout.write("\n")
    sys.stdout.flush()


def _write_stream_delta(delta: str) -> None:
    _write_json({"type": "delta", "delta": delta})


def _write_stream_audio(audio: dict[str, Any]) -> None:
    _write_json({"type": "audio", "audio": audio})


def _gui_log_path(root: Path | str | None = None) -> Path:
    override = os.environ.get(GUI_LOG_ENV)
    if override:
        return Path(override)

    base = Path(root) if root is not None else ROOT
    return base / "logs" / "chat_gui_bridge.log"


def _gui_log(root: Path | str | None, message: str) -> None:
    try:
        path = _gui_log_path(root)
        path.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
        with path.open("a", encoding="utf-8") as file:
            file.write(f"{timestamp} pid={os.getpid()} {message}\n")
    except OSError:
        pass


def _format_exception(exc: Exception) -> str:
    return f"type={type(exc).__name__} message={_safe_log_text(str(exc))}"


def _safe_log_text(value: str, max_length: int = 1000) -> str:
    text = value.replace("\r", "\\r").replace("\n", "\\n")
    if len(text) > max_length:
        return text[:max_length] + "...[truncated]"
    return text


def _payload_log_summary(payload: dict[str, Any]) -> str:
    parts = []
    if payload.get("session_file"):
        parts.append(f"session_file={Path(str(payload['session_file'])).name}")
    if payload.get("session_ref"):
        parts.append(f"session_ref={_safe_log_text(str(payload['session_ref']))}")
    if payload.get("request_id"):
        parts.append(f"request_id={_safe_log_text(str(payload['request_id']))}")
    if "content" in payload:
        parts.append(f"content_chars={len(str(payload.get('content') or ''))}")
    if "no_ai" in payload:
        parts.append(f"no_ai={bool(payload.get('no_ai'))}")
    if isinstance(payload.get("selected_ids"), list):
        parts.append(f"selected_ids={len(payload['selected_ids'])}")
    return ",".join(parts) if parts else "-"


def _result_log_summary(result: dict[str, Any]) -> str:
    if "session" in result and isinstance(result["session"], dict):
        session = result["session"]
        return f"session={session.get('session_id', '-')}"
    if "sessions" in result and isinstance(result["sessions"], list):
        return f"sessions={len(result['sessions'])}"
    if "request_id" in result:
        return f"request_id={result['request_id']}"
    if "raw_file" in result:
        return f"raw_file={result['raw_file']}"
    if "imported_count" in result:
        return f"imported_count={result['imported_count']}"
    if "total_count" in result:
        return f"total_count={result['total_count']}"
    return "-"


def _open_folder(path: Path) -> None:
    if os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
        return

    command = ["open", str(path)] if sys.platform == "darwin" else ["xdg-open", str(path)]
    subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def _new_job_id() -> str:
    return uuid4().hex


def _required_job_id(value: Any) -> str:
    text = str(value or "").strip()
    if not JOB_ID_PATTERN.fullmatch(text):
        raise ValueError("job_id が不正です。")
    return text


def _job_dir(root: Path) -> Path:
    return root / "logs" / "chat_gui_jobs"


def _job_status_path(root: Path, job_id: str) -> Path:
    return _job_dir(root) / f"{job_id}.json"


def _job_log_path(root: Path, job_id: str) -> Path:
    return _job_dir(root) / f"{job_id}.log"


def _job_event_log(root: Path, job_id: str, message: str) -> None:
    try:
        path = _job_log_path(root, job_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        timestamp = _now_iso()
        with path.open("a", encoding="utf-8") as file:
            file.write(f"{timestamp} pid={os.getpid()} {message}\n")
    except OSError:
        pass


def _job_worker_event(job_id: str, message: str) -> None:
    timestamp = _now_iso()
    sys.stderr.write(f"{timestamp} pid={os.getpid()} job_id={job_id} {message}\n")
    sys.stderr.flush()


def _job_cancel_file(root: Path, job_id: str) -> Path:
    return _job_dir(root) / f"{job_id}.cancel"


def _organize_sessions_lock_path(root: Path) -> Path:
    return _job_dir(root) / "organize-sessions.lock"


def _claim_organize_sessions_lock(root: Path, job_id: str) -> dict[str, Any] | None:
    lock_path = _organize_sessions_lock_path(root)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_value = {"job_id": job_id, "created_at": _now_iso()}

    for _ in range(120):
        try:
            with lock_path.open("x", encoding="utf-8") as file:
                json.dump(lock_value, file, ensure_ascii=False, indent=2)
                file.write("\n")
            return None
        except FileExistsError:
            existing_lock = _read_finalize_lock(lock_path)
            existing_job_id = str(existing_lock.get("job_id") or "") if existing_lock else ""
            if JOB_ID_PATTERN.fullmatch(existing_job_id):
                try:
                    existing = _refresh_organize_sessions_job_status(root, existing_job_id)
                except (FileNotFoundError, ValueError):
                    existing = None
                if existing is not None and existing.get("status") in FINALIZE_ACTIVE_STATUSES:
                    return existing
                if existing is not None or not _lock_is_recent(existing_lock, lock_path):
                    _release_organize_sessions_lock(root=root, job_id=existing_job_id)
                    continue
            elif not _lock_is_recent(existing_lock, lock_path):
                lock_path.unlink(missing_ok=True)
                continue
            time.sleep(0.05)

    raise RuntimeError("データ整理ジョブのロックを取得できませんでした。少し待ってから再実行してください。")


def _release_organize_sessions_lock(root: Path, job_id: str) -> None:
    lock_path = _organize_sessions_lock_path(root)
    if not lock_path.exists():
        return
    lock_value = _read_finalize_lock(lock_path)
    if str(lock_value.get("job_id") or "") != job_id:
        return
    lock_path.unlink(missing_ok=True)


def _finalize_lock_path(root: Path, session_file: Path) -> Path:
    session_key = str(session_file.resolve())
    if os.name == "nt":
        session_key = session_key.casefold()
    digest = hashlib.sha256(session_key.encode("utf-8")).hexdigest()[:24]
    return _job_dir(root) / f"session-{digest}.lock"


def _claim_finalize_lock(root: Path, session_file: Path, job_id: str) -> dict[str, Any] | None:
    lock_path = _finalize_lock_path(root, session_file)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_value = {
        "job_id": job_id,
        "session_file": _display_path(session_file, root),
        "created_at": _now_iso(),
    }

    for _ in range(120):
        try:
            with lock_path.open("x", encoding="utf-8") as file:
                json.dump(lock_value, file, ensure_ascii=False, indent=2)
                file.write("\n")
            return None
        except FileExistsError:
            existing_lock = _read_finalize_lock(lock_path)
            existing_job_id = str(existing_lock.get("job_id") or "") if existing_lock else ""
            if JOB_ID_PATTERN.fullmatch(existing_job_id):
                try:
                    existing = _refresh_finalize_job_status(root, existing_job_id)
                except (FileNotFoundError, ValueError):
                    existing = None
                if existing is not None and existing.get("status") in FINALIZE_ACTIVE_STATUSES:
                    if _job_matches_session(root, existing, session_file):
                        return existing
                if existing is not None or not _lock_is_recent(existing_lock, lock_path):
                    _release_finalize_lock(root=root, session_file=session_file, job_id=existing_job_id)
                    continue
            elif not _lock_is_recent(existing_lock, lock_path):
                lock_path.unlink(missing_ok=True)
                continue
            time.sleep(0.05)

    raise RuntimeError("整理ジョブのセッションロックを取得できませんでした。少し待ってから再実行してください。")


def _read_finalize_lock(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}


def _lock_is_recent(value: dict[str, Any], path: Path) -> bool:
    created_age = _seconds_since(value.get("created_at"))
    if created_age != float("inf"):
        return created_age < FINALIZE_WORKER_START_GRACE_SECONDS
    try:
        return time.time() - path.stat().st_mtime < FINALIZE_WORKER_START_GRACE_SECONDS
    except OSError:
        return False


def _release_finalize_lock(root: Path, session_file: Path, job_id: str) -> None:
    lock_path = _finalize_lock_path(root, session_file)
    if not lock_path.exists():
        return
    lock_value = _read_finalize_lock(lock_path)
    if str(lock_value.get("job_id") or "") != job_id:
        return
    lock_path.unlink(missing_ok=True)


def _find_active_finalize_job(root: Path, session_file: Path) -> dict[str, Any] | None:
    job_dir = _job_dir(root)
    if not job_dir.exists():
        return None

    active: list[dict[str, Any]] = []
    for path in job_dir.glob("*.json"):
        try:
            value = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(value, dict) or not _job_matches_session(root, value, session_file):
            continue
        job_id = str(value.get("job_id") or "")
        if not JOB_ID_PATTERN.fullmatch(job_id):
            continue
        try:
            refreshed = _refresh_finalize_job_status(root, job_id)
        except (FileNotFoundError, ValueError):
            continue
        if refreshed.get("status") in FINALIZE_ACTIVE_STATUSES:
            active.append(refreshed)

    if not active:
        return None

    active.sort(key=lambda item: str(item.get("created_at") or ""))
    primary = active[0]
    for duplicate in active[1:]:
        duplicate_id = str(duplicate["job_id"])
        cancel_file = _job_cancel_file(root, duplicate_id)
        cancel_file.parent.mkdir(parents=True, exist_ok=True)
        cancel_file.write_text(_now_iso(), encoding="utf-8")
        _write_job_status(
            root,
            duplicate_id,
            {
                "message": "同じセッションの整理ジョブが既に動いているため、この重複ジョブを停止します。",
                "cancel_requested_at": _now_iso(),
            },
        )
        _gui_log(root, f"finalize_job.duplicate_cancelled job_id={duplicate_id} primary={primary['job_id']}")
    return primary


def _find_any_active_finalize_job(root: Path) -> dict[str, Any] | None:
    job_dir = _job_dir(root)
    if not job_dir.exists():
        return None

    for path in job_dir.glob("*.json"):
        try:
            status = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(status, dict) or status.get("name") != "finalize-session":
            continue
        job_id = str(status.get("job_id") or "")
        if not JOB_ID_PATTERN.fullmatch(job_id):
            continue
        refreshed = _refresh_finalize_job_status(root, job_id)
        if refreshed.get("status") in FINALIZE_ACTIVE_STATUSES:
            return refreshed
    return None


def _find_active_organize_sessions_job(root: Path) -> dict[str, Any] | None:
    job_dir = _job_dir(root)
    if not job_dir.exists():
        return None

    active: list[dict[str, Any]] = []
    for path in job_dir.glob("*.json"):
        try:
            status = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(status, dict) or status.get("name") != ORGANIZE_SESSIONS_JOB_NAME:
            continue
        job_id = str(status.get("job_id") or "")
        if not JOB_ID_PATTERN.fullmatch(job_id):
            continue
        refreshed = _refresh_organize_sessions_job_status(root, job_id)
        if refreshed.get("status") in FINALIZE_ACTIVE_STATUSES:
            active.append(refreshed)

    if not active:
        return None
    return min(active, key=lambda item: str(item.get("created_at") or ""))


def _organize_session_targets(root: Path) -> list[Path]:
    sessions = list_resumable_sessions(root=root, limit=None)
    targets = [
        session
        for session in sessions
        if not load_session_personalization(root, session.jsonl_file).exclude_from_memory
        and get_session_organization(root=root, session_file=session.jsonl_file).get("can_organize")
    ]
    targets.sort(key=lambda session: session.last_user_at)
    return [session.jsonl_file for session in targets]


def _job_matches_session(root: Path, status: dict[str, Any], session_file: Path) -> bool:
    value = status.get("session_file")
    if not value:
        return False
    path = Path(str(value))
    if not path.is_absolute():
        path = root / path
    try:
        return path.resolve() == session_file.resolve()
    except OSError:
        return False


def _read_job_status(root: Path, job_id: str) -> dict[str, Any]:
    path = _job_status_path(root, job_id)
    if not path.exists():
        raise FileNotFoundError(f"ジョブが見つかりません: {job_id}")

    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"ジョブ状態ファイルが壊れています: {job_id}") from exc
    if not isinstance(value, dict):
        raise ValueError(f"ジョブ状態ファイルが不正です: {job_id}")
    return value


def _refresh_finalize_job_status(root: Path, job_id: str) -> dict[str, Any]:
    status = _read_job_status(root, job_id)
    if status.get("status") not in FINALIZE_ACTIVE_STATUSES:
        return status

    worker_pid = status.get("worker_pid")
    try:
        pid = int(worker_pid) if worker_pid is not None else None
    except (TypeError, ValueError):
        pid = None

    launcher_pid_value = status.get("launcher_pid")
    try:
        launcher_pid = int(launcher_pid_value) if launcher_pid_value is not None else None
    except (TypeError, ValueError):
        launcher_pid = None

    worker_missing = pid is not None and not _process_exists(pid)
    launcher_alive = launcher_pid is not None and _process_exists(launcher_pid)
    worker_never_started = (
        pid is None
        and not launcher_alive
        and _seconds_since(status.get("created_at")) >= FINALIZE_WORKER_START_GRACE_SECONDS
    )
    if not worker_missing and not worker_never_started:
        return status

    status = _write_job_status(
        root,
        job_id,
        {
            "status": "failed",
            "stage": "failed",
            "message": "整理ワーカーが終了したため、ジョブを失敗状態へ回収しました。再実行できます。",
            "error": "Finalize worker exited before recording a terminal result.",
            "finished_at": _now_iso(),
        },
    )
    _job_event_log(root, job_id, f"worker.missing worker_pid={pid or '-'} status=failed")
    session_value = status.get("session_file")
    if session_value:
        session_path = Path(str(session_value))
        if not session_path.is_absolute():
            session_path = root / session_path
        _release_finalize_lock(root=root, session_file=session_path, job_id=job_id)
    _gui_log(root, f"finalize_job.orphan_recovered job_id={job_id} worker_pid={pid or '-'}")
    return status


def _refresh_organize_sessions_job_status(root: Path, job_id: str) -> dict[str, Any]:
    status = _read_job_status(root, job_id)
    if status.get("name") != ORGANIZE_SESSIONS_JOB_NAME:
        raise ValueError(f"データ整理ジョブではありません: {job_id}")
    if status.get("status") not in FINALIZE_ACTIVE_STATUSES:
        return status

    worker_pid = status.get("worker_pid")
    try:
        pid = int(worker_pid) if worker_pid is not None else None
    except (TypeError, ValueError):
        pid = None

    launcher_pid_value = status.get("launcher_pid")
    try:
        launcher_pid = int(launcher_pid_value) if launcher_pid_value is not None else None
    except (TypeError, ValueError):
        launcher_pid = None

    worker_missing = pid is not None and not _process_exists(pid)
    launcher_alive = launcher_pid is not None and _process_exists(launcher_pid)
    worker_never_started = (
        pid is None
        and not launcher_alive
        and _seconds_since(status.get("created_at")) >= FINALIZE_WORKER_START_GRACE_SECONDS
    )
    if not worker_missing and not worker_never_started:
        return status

    status = _write_job_status(
        root,
        job_id,
        {
            "status": "failed",
            "stage": "failed",
            "message": "データ整理ワーカーが終了したため、ジョブを失敗状態へ回収しました。再実行できます。",
            "error": "Organize sessions worker exited before recording a terminal result.",
            "finished_at": _now_iso(),
            "current_session": None,
        },
    )
    _job_event_log(root, job_id, f"worker.missing worker_pid={pid or '-'} status=failed")
    _release_organize_sessions_lock(root=root, job_id=job_id)
    _gui_log(root, f"organize_sessions.orphan_recovered job_id={job_id} worker_pid={pid or '-'}")
    return status


def _seconds_since(value: Any) -> float:
    try:
        then = datetime.fromisoformat(str(value))
        now = datetime.now().astimezone()
        if then.tzinfo is None:
            then = then.replace(tzinfo=now.tzinfo)
        return max(0.0, (now - then).total_seconds())
    except (TypeError, ValueError):
        return float("inf")


def _process_exists(pid: int) -> bool:
    if pid <= 0:
        return False
    if pid == os.getpid():
        return True

    if os.name == "nt":
        try:
            import ctypes

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            kernel32.OpenProcess.argtypes = [ctypes.c_uint32, ctypes.c_int, ctypes.c_uint32]
            kernel32.OpenProcess.restype = ctypes.c_void_p
            kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
            kernel32.CloseHandle.restype = ctypes.c_int
            handle = kernel32.OpenProcess(0x1000, False, pid)
            if handle:
                kernel32.CloseHandle(handle)
                return True
            return ctypes.get_last_error() == 5
        except (AttributeError, OSError):
            return False

    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    return True


def _write_job_status(root: Path, job_id: str, update: dict[str, Any]) -> dict[str, Any]:
    path = _job_status_path(root, job_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    current: dict[str, Any] = {}
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            current = loaded if isinstance(loaded, dict) else {}
        except json.JSONDecodeError:
            current = {}
    current.update(update)
    temp_path = path.with_name(f"{path.name}.{os.getpid()}.{uuid4().hex}.tmp")
    temp_path.write_text(json.dumps(current, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)
    return current


def _spawn_finalize_worker(root: Path, payload: dict[str, Any], log_path: Path) -> subprocess.Popen[str]:
    return _spawn_background_worker(root=root, payload=payload, log_path=log_path, worker_command="run-finalize-job")


def _spawn_organize_sessions_worker(root: Path, payload: dict[str, Any], log_path: Path) -> subprocess.Popen[str]:
    return _spawn_background_worker(root=root, payload=payload, log_path=log_path, worker_command="run-organize-sessions-job")


def _spawn_background_worker(
    root: Path,
    payload: dict[str, Any],
    log_path: Path,
    worker_command: str,
) -> subprocess.Popen[str]:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_handle = log_path.open("a", encoding="utf-8")
    creationflags = _finalize_worker_creationflags()
    try:
        command = [sys.executable, str(Path(__file__).resolve()), worker_command]
        popen_kwargs = {
            "cwd": root,
            "env": {**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
            "stdin": subprocess.PIPE,
            "stdout": log_handle,
            "stderr": subprocess.STDOUT,
            "text": True,
            "encoding": "utf-8",
            "creationflags": creationflags,
            "close_fds": True,
        }
        try:
            process = subprocess.Popen(command, **popen_kwargs)
        except OSError as exc:
            breakaway_flag = getattr(subprocess, "CREATE_BREAKAWAY_FROM_JOB", 0)
            if os.name != "nt" or not breakaway_flag or not (creationflags & breakaway_flag):
                raise
            fallback_flags = creationflags & ~breakaway_flag
            log_handle.write(
                f"{_now_iso()} pid={os.getpid()} worker.detach_fallback type={type(exc).__name__} "
                f"message={_safe_log_text(str(exc))}\n"
            )
            log_handle.flush()
            popen_kwargs["creationflags"] = fallback_flags
            process = subprocess.Popen(command, **popen_kwargs)
        assert process.stdin is not None
        process.stdin.write(json.dumps(payload, ensure_ascii=False))
        process.stdin.close()
        BACKGROUND_PROCESSES.append(process)
        log_handle.close()
        return process
    except Exception:
        log_handle.close()
        raise


def _finalize_worker_creationflags() -> int:
    if os.name != "nt":
        return 0
    return (
        getattr(subprocess, "DETACHED_PROCESS", 0)
        | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
        | getattr(subprocess, "CREATE_BREAKAWAY_FROM_JOB", 0)
    )


def _run_finalize_job(root: Path, payload: dict[str, Any], job_id: str) -> None:
    session_file = _resolve_existing_session(root=root, value=payload.get("session_file"))
    cancel_file = _job_cancel_file(root, job_id)
    _write_job_status(
        root,
        job_id,
        {
            "status": "running",
            "stage": "starting",
            "message": "整理処理を開始しました。",
            "percent": 5,
            "started_at": _now_iso(),
            "worker_pid": os.getpid(),
        },
    )
    _job_worker_event(job_id, f"worker.start session={session_file.name}")

    def progress(percent: int, message: str) -> None:
        if cancel_file.exists():
            raise AssistantGenerationCancelled("整理ジョブをキャンセルしました。")
        _write_job_status(
            root,
            job_id,
            {
                "status": "running",
                "stage": _stage_from_progress_message(message),
                "message": message,
                "percent": percent,
            },
        )
        _job_worker_event(
            job_id,
            f"worker.progress percent={percent} stage={_stage_from_progress_message(message)} message={_safe_log_text(message)}",
        )

    try:
        _ensure_session_can_be_organized(root, session_file)
        run_command = _cancelable_run_command(root=root, cancel_file=cancel_file)
        result = finalize_live_chat(
            root=root,
            session_file=session_file,
            run_codex=bool(payload.get("run_codex", True)),
            commit=False,
            force=True,
            codex_command=str(payload.get("codex_command") or "codex.cmd"),
            codex_sandbox="workspace-write",
            codex_approval=str(payload.get("codex_approval") or "never"),
            progress=progress,
            run_command=run_command,
        )
        _write_job_status(
            root,
            job_id,
            {
                "status": "succeeded",
                "stage": "done",
                "message": "整理処理が完了しました。",
                "percent": 100,
                "finished_at": _now_iso(),
                "result": {
                    "ok": True,
                    "session": _serialize_session_file(result.jsonl_file, root),
                    "jsonl_file": _display_path(result.jsonl_file, root),
                    "raw_file": _display_path(result.raw_file, root),
                    "task_file": _display_path(result.task_file, root),
                    "imported_at": result.imported_at.isoformat(timespec="seconds"),
                    "codex_updated": bool(result.codex),
                    "git_committed": bool(result.git and result.git.committed),
                    "organization": result.organization,
                },
            },
        )
        _job_worker_event(job_id, "worker.succeeded percent=100")
    except AssistantGenerationCancelled as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "cancelled",
                "stage": "cancelled",
                "message": str(exc),
                "finished_at": _now_iso(),
            },
        )
        _job_worker_event(job_id, f"worker.cancelled message={_safe_log_text(str(exc))}")
    except Exception as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "failed",
                "stage": "failed",
                "message": "整理処理に失敗しました。",
                "error": f"{type(exc).__name__}: {exc}",
                "finished_at": _now_iso(),
            },
        )
        _job_worker_event(job_id, f"worker.failed {_format_exception(exc)}")
        raise
    finally:
        _release_finalize_lock(root=root, session_file=session_file, job_id=job_id)


def _run_organize_sessions_job(root: Path, payload: dict[str, Any], job_id: str) -> None:
    raw_targets = payload.get("session_files")
    if not isinstance(raw_targets, list):
        raise ValueError("データ整理対象が不正です。")
    session_values = [str(value) for value in raw_targets if str(value).strip()]
    if not session_values:
        raise ValueError("データ整理対象がありません。")

    cancel_file = _job_cancel_file(root, job_id)
    total = len(session_values)
    completed: list[str] = []
    skipped: list[str] = []
    failed: list[dict[str, str]] = []
    _write_job_status(
        root,
        job_id,
        {
            "status": "running",
            "stage": "starting",
            "message": f"データ整理を開始しました（{total}件）。",
            "percent": 1,
            "started_at": _now_iso(),
            "worker_pid": os.getpid(),
        },
    )
    _job_worker_event(job_id, f"worker.start targets={total}")

    def update_progress(index: int, session_id: str, percent: int, message: str) -> None:
        if cancel_file.exists():
            raise AssistantGenerationCancelled("データ整理をキャンセルしました。")
        overall_percent = min(99, max(1, int((((index - 1) + (percent / 100)) / total) * 100)))
        _write_job_status(
            root,
            job_id,
            {
                "status": "running",
                "stage": _stage_from_progress_message(message),
                "message": f"{index}/{total}: {session_id} — {message}",
                "percent": overall_percent,
                "current_session": session_id,
                "completed_count": len(completed),
                "failed_count": len(failed),
                "skipped_count": len(skipped),
                "failed_sessions": failed,
            },
        )
        _job_worker_event(
            job_id,
            f"worker.progress index={index}/{total} session={session_id} percent={overall_percent} "
            f"stage={_stage_from_progress_message(message)} message={_safe_log_text(message)}",
        )

    try:
        run_command = _cancelable_run_command(root=root, cancel_file=cancel_file)
        for index, session_value in enumerate(session_values, start=1):
            if cancel_file.exists():
                raise AssistantGenerationCancelled("データ整理をキャンセルしました。")

            try:
                session_file = _resolve_existing_session(root=root, value=session_value)
            except FileNotFoundError:
                skipped.append(Path(session_value).stem)
                continue

            if load_session_personalization(root, session_file).exclude_from_memory:
                skipped.append(session_file.stem)
                update_progress(index, session_file.stem, 100, "一時チャットのため整理対象から除外しました。")
                continue

            organization = get_session_organization(root=root, session_file=session_file)
            if not organization.get("can_organize"):
                skipped.append(session_file.stem)
                update_progress(index, session_file.stem, 100, "既に整理済みのためスキップしました。")
                continue

            update_progress(index, session_file.stem, 1, "整理を開始しています。")
            try:
                finalize_live_chat(
                    root=root,
                    session_file=session_file,
                    run_codex=bool(payload.get("run_codex", True)),
                    commit=False,
                    force=True,
                    codex_command=str(payload.get("codex_command") or "codex.cmd"),
                    codex_sandbox="workspace-write",
                    codex_approval=str(payload.get("codex_approval") or "never"),
                    progress=lambda percent, message, current_index=index, current_session=session_file.stem: update_progress(
                        current_index,
                        current_session,
                        percent,
                        message,
                    ),
                    run_command=run_command,
                )
            except AssistantGenerationCancelled:
                raise
            except Exception as exc:
                failed.append({"session_id": session_file.stem, "error": f"{type(exc).__name__}: {exc}"})
                _job_worker_event(job_id, f"worker.session_failed session={session_file.stem} {_format_exception(exc)}")
            else:
                completed.append(session_file.stem)

            _write_job_status(
                root,
                job_id,
                {
                    "status": "running",
                    "stage": "running",
                    "message": f"{index}/{total}件を処理しました。",
                    "percent": min(99, int((index / total) * 100)),
                    "current_session": None,
                    "completed_count": len(completed),
                    "failed_count": len(failed),
                    "skipped_count": len(skipped),
                    "failed_sessions": failed,
                },
            )

        result = {
            "total_sessions": total,
            "completed_sessions": completed,
            "failed_sessions": failed,
            "skipped_sessions": skipped,
        }
        if failed:
            _write_job_status(
                root,
                job_id,
                {
                    "status": "failed",
                    "stage": "done",
                    "message": f"データ整理を完了しましたが、{len(failed)}件で失敗しました。",
                    "error": "一部のセッションを整理できませんでした。詳細はジョブログを確認してください。",
                    "percent": 100,
                    "current_session": None,
                    "completed_count": len(completed),
                    "failed_count": len(failed),
                    "skipped_count": len(skipped),
                    "failed_sessions": failed,
                    "finished_at": _now_iso(),
                    "result": result,
                },
            )
            _job_worker_event(job_id, f"worker.completed_with_failures completed={len(completed)} failed={len(failed)}")
        else:
            _write_job_status(
                root,
                job_id,
                {
                    "status": "succeeded",
                    "stage": "done",
                    "message": f"データ整理が完了しました（{len(completed)}件、スキップ{len(skipped)}件）。",
                    "percent": 100,
                    "current_session": None,
                    "completed_count": len(completed),
                    "failed_count": 0,
                    "skipped_count": len(skipped),
                    "failed_sessions": [],
                    "finished_at": _now_iso(),
                    "result": result,
                },
            )
            _job_worker_event(job_id, f"worker.succeeded completed={len(completed)} skipped={len(skipped)}")
    except AssistantGenerationCancelled as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "cancelled",
                "stage": "cancelled",
                "message": str(exc),
                "current_session": None,
                "completed_count": len(completed),
                "failed_count": len(failed),
                "skipped_count": len(skipped),
                "failed_sessions": failed,
                "finished_at": _now_iso(),
                "result": {
                    "total_sessions": total,
                    "completed_sessions": completed,
                    "failed_sessions": failed,
                    "skipped_sessions": skipped,
                },
            },
        )
        _job_worker_event(job_id, f"worker.cancelled completed={len(completed)} failed={len(failed)}")
    except Exception as exc:
        _write_job_status(
            root,
            job_id,
            {
                "status": "failed",
                "stage": "failed",
                "message": "データ整理に失敗しました。",
                "error": f"{type(exc).__name__}: {exc}",
                "current_session": None,
                "completed_count": len(completed),
                "failed_count": len(failed),
                "skipped_count": len(skipped),
                "failed_sessions": failed,
                "finished_at": _now_iso(),
            },
        )
        _job_worker_event(job_id, f"worker.failed {_format_exception(exc)}")
        raise
    finally:
        _release_organize_sessions_lock(root=root, job_id=job_id)


def _stage_from_progress_message(message: str) -> str:
    lowered = message.lower()
    if "raw" in lowered:
        return "raw"
    if "memory" in lowered or "journal" in lowered or "summary" in lowered:
        return "memory"
    if "index" in lowered:
        return "index"
    return "running"


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _payload_root(payload: dict[str, Any]) -> Path:
    root = payload.get("root")
    return Path(root) if root else ROOT


def _personalization_update_fields(value: Any, *, allowed: set[str], label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{label} はJSON objectで指定してください。")
    unknown = set(value) - allowed
    if unknown:
        raise ValueError(f"{label} に未対応の項目があります: {', '.join(sorted(unknown))}")
    if not value:
        raise ValueError(f"{label} に更新項目を指定してください。")
    fields = {key: value[key] for key in allowed if key in value}
    for name in {"memory_enabled", "past_chat_search_enabled", "temporary"} & set(fields):
        if not isinstance(fields[name], bool):
            raise ValueError(f"{name} は true または false で指定してください。")
    if "project_scope" in fields:
        fields["project_scope"] = validate_project_scope(fields["project_scope"])
    return fields


def _ensure_session_can_be_organized(root: Path, session_file: Path) -> None:
    if load_session_personalization(root, session_file).exclude_from_memory:
        raise ValueError("一時チャットは記憶整理の対象外です。通常チャットとして新しいセッションを作成してください。")


def _read_aloud_request_id(value: Any) -> str:
    request_id = str(value or "").strip()
    if not READ_ALOUD_REQUEST_ID_PATTERN.fullmatch(request_id):
        raise ValueError("読み上げ request_id が不正です。")
    return request_id


def _read_aloud_runtime_dir() -> Path:
    return temporary_audio_dir().resolve()


def _read_aloud_cancel_file(root: Path, request_id: str) -> Path:
    del root
    return _read_aloud_runtime_dir() / f"read-aloud-{request_id}.cancel"


def _read_aloud_audio_path(root: Path, value: Any) -> Path:
    raw_path = str(value or "").strip()
    if not raw_path:
        raise ValueError("読み上げ音声ファイルが指定されていません。")

    candidate = Path(raw_path)
    if not candidate.is_absolute():
        candidate = root / candidate
    try:
        resolved = candidate.resolve(strict=False)
        resolved.relative_to(_read_aloud_runtime_dir())
    except (OSError, ValueError) as exc:
        raise ValueError("読み上げ用の一時音声ファイルだけを削除できます。") from exc
    if resolved.suffix.lower() != ".wav" or not resolved.name.startswith("read-aloud-"):
        raise ValueError("読み上げ用の一時音声ファイルだけを削除できます。")
    return resolved


def _tts_log(root: Path, message: str) -> None:
    try:
        path = root / "logs" / "tts" / "kokoro_tts.log"
        path.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
        with path.open("a", encoding="utf-8") as file:
            file.write(f"{timestamp} pid={os.getpid()} {message}\n")
    except OSError:
        pass


def _resolve_chatgpt_export_source(value: Any) -> Path:
    text = str(value or "").strip()
    if not text:
        raise ValueError(
            "ChatGPTエクスポートのフォルダ、zip、conversations.json、または conversations-*.json を選択してください。"
        )

    try:
        path = Path(text).expanduser().resolve(strict=True)
    except OSError as exc:
        raise FileNotFoundError("選択したChatGPTエクスポートが見つかりません。") from exc

    if path.is_dir() or path.suffix.lower() == ".zip" or CONVERSATION_FILE_PATTERN.fullmatch(path.name):
        return path
    raise ValueError(
        "ChatGPTエクスポートのフォルダ、zip、conversations.json、または conversations-*.json を選択してください。"
    )


def _load_chatgpt_export_source(source_path: Path):
    try:
        return load_export(source_path)
    except OSError as exc:
        raise ValueError("ChatGPTエクスポートを読み取れません。ファイルのアクセス権を確認してください。") from exc


def _selected_chatgpt_import_ids(value: Any) -> set[str]:
    if not isinstance(value, list):
        raise ValueError("取り込み対象の会話を選択してください。")
    selected_ids = {str(item).strip() for item in value if str(item).strip()}
    if not selected_ids:
        raise ValueError("取り込み対象の会話を少なくとも1件選択してください。")
    return selected_ids


def _serialize_chatgpt_import_conversation(
    conversation: ExportConversation,
    *,
    import_state: str,
) -> dict[str, Any]:
    return {
        "source_id": conversation.source_id,
        "title": conversation.title,
        "created_at": _chatgpt_import_timestamp(conversation.created_at),
        "updated_at": _chatgpt_import_timestamp(conversation.updated_at),
        "message_count": len(conversation.messages),
        "duplicate": import_state == IMPORT_STATE_DUPLICATE,
        "import_state": import_state,
        "source_message_count": conversation.source_message_count,
        "skipped_message_count": conversation.skipped_message_count,
        "non_text_message_count": conversation.non_text_message_count,
        "attachment_count": conversation.attachment_count,
        "non_text_part_count": conversation.non_text_part_count,
        "audio_transcription_count": conversation.audio_transcription_count,
        "empty_conversation": conversation.empty_conversation,
    }


def _chatgpt_import_timestamp(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _optional_request_id(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return _sanitize_request_id(text)


def _required_request_id(value: Any) -> str:
    request_id = _optional_request_id(value)
    if not request_id:
        raise ValueError("request_id is required.")
    return request_id


def _sanitize_request_id(value: str) -> str:
    if len(value) > 128:
        raise ValueError("request_id is too long.")
    if not REQUEST_ID_PATTERN.fullmatch(value):
        raise ValueError("request_id contains invalid characters.")
    return value


def _cancel_file_path(root: Path, request_id: str) -> Path:
    return root / "logs" / "chat_gui_cancel" / f"{request_id}.cancel"


def _clear_cancel_file(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass
    except OSError:
        pass


def _cancelable_run_command(root: Path, cancel_file: Path) -> Callable[..., subprocess.CompletedProcess[str]]:
    def run_command(command: list[str], **kwargs: Any) -> subprocess.CompletedProcess[str]:
        if cancel_file.exists():
            raise AssistantGenerationCancelled("返答生成を停止しました。")

        creationflags = _cancelable_subprocess_creationflags()

        process = subprocess.Popen(
            command,
            cwd=kwargs.get("cwd", root),
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=bool(kwargs.get("text", False)),
            encoding=kwargs.get("encoding"),
            creationflags=creationflags,
        )

        input_text = kwargs.get("input")
        try:
            try:
                stdout, stderr = process.communicate(input=input_text, timeout=0.25)
            except subprocess.TimeoutExpired:
                while True:
                    if cancel_file.exists():
                        _terminate_process_tree(process)
                        _drain_cancelled_process(process)
                        raise AssistantGenerationCancelled("返答生成を停止しました。")
                    try:
                        stdout, stderr = process.communicate(timeout=0.25)
                        break
                    except subprocess.TimeoutExpired:
                        continue
        except FileNotFoundError:
            raise

        if cancel_file.exists():
            raise AssistantGenerationCancelled("返答生成を停止しました。")

        return subprocess.CompletedProcess(command, process.returncode, stdout, stderr)

    return run_command


def _cancelable_subprocess_creationflags() -> int:
    if os.name != "nt":
        return 0
    return getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0) | getattr(subprocess, "CREATE_NO_WINDOW", 0)


def _terminate_process_tree(process: subprocess.Popen[Any]) -> None:
    if process.poll() is not None:
        return

    if os.name == "nt":
        try:
            subprocess.run(
                ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
            )
            return
        except OSError:
            pass

    process.kill()


def _drain_cancelled_process(process: subprocess.Popen[Any]) -> None:
    try:
        process.communicate(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()
        process.communicate()


def _resolve_or_create_session(root: Path, value: Any) -> Path:
    if value:
        path = _resolve_session_path(root=root, value=value, must_exist=False)
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    session = create_live_session(root=root)
    settings = load_personalization_settings(root)
    update_session_personalization(
        root,
        session.path,
        temporary=False,
        memory_enabled=settings.memory_enabled,
        past_chat_search_enabled=settings.past_chat_search_enabled,
        project_scope=settings.project_scope,
    )
    return session.path


def _resolve_existing_session(root: Path, value: Any) -> Path:
    return _resolve_session_path(root=root, value=value, must_exist=True)


def _resolve_session_path(root: Path, value: Any, must_exist: bool) -> Path:
    if not value:
        raise ValueError("session_file is required.")

    path = Path(str(value))
    if not path.is_absolute():
        path = root / path

    path = path.resolve()
    try:
        path.relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError("session_file must be inside the AI-LifeOS root.") from exc

    if path.suffix != ".jsonl":
        raise ValueError("session_file must be a .jsonl file.")
    if must_exist and not path.exists():
        raise FileNotFoundError(f"session_file not found: {path}")

    return path


def _read_live_messages(path: Path) -> list[LiveMessage]:
    if not path.exists():
        return []

    messages: list[LiveMessage] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSONL at line {line_number}: {path}") from exc
        messages.append(_message_from_record(record))
    return messages


def _message_from_record(record: dict[str, Any]) -> LiveMessage:
    return LiveMessage(
        role=str(record["role"]),
        content=str(record["content"]),
        timestamp=_parse_datetime(str(record["timestamp"])),
    )


def _parse_datetime(value: str) -> datetime:
    return datetime.fromisoformat(value)


def _session_started_at(path: Path) -> datetime:
    messages = _read_live_messages(path)
    if messages:
        return messages[0].timestamp
    return datetime.now().astimezone()


def _save_session_metadata(root: Path, session_file: Path, status: str) -> None:
    try:
        save_session(root=root, session_file=session_file, status=status)
    except Exception as exc:
        _gui_log(root, f"save_session_metadata.error session={session_file.stem} {_format_exception(exc)}")


def _serialize_session_file(path: Path, root: Path) -> dict[str, Any]:
    return {
        "session_id": path.stem,
        "jsonl_file": _display_path(path, root),
        "organization": _serialize_organization(root=root, session_file=path),
        "personalization": serialize_personalization(load_session_personalization(root, path)),
    }


def _rebuild_chatgpt_import_index(root: Path) -> tuple[bool, str, str | None]:
    """Refresh the derived index without changing the outcome of a completed import."""

    try:
        rebuild_index(root=root)
        health = inspect_index_health(root=root)
    except Exception as exc:
        # Importing raw.md is the durable operation.  Index failures must not
        # make that successful write look rolled back, and exception text may
        # contain an absolute path or private source detail.
        _gui_log(root, f"chatgpt_import.index_failed error_type={type(exc).__name__}")
        return (
            False,
            "error",
            "会話の取り込みは完了しましたが、検索indexの更新に失敗しました。後で再構築してください。",
        )

    if health.status != "fresh":
        status = _safe_log_text(str(health.status), max_length=40) or "unknown"
        _gui_log(root, f"chatgpt_import.index_incomplete status={status}")
        return (
            False,
            status,
            "会話の取り込みは完了しましたが、検索indexの最新状態を確認できませんでした。後で再構築してください。",
        )

    _gui_log(
        root,
        f"chatgpt_import.index_updated status=fresh sources={health.source_count}",
    )
    return True, "fresh", None


def _serialize_message(message: LiveMessage) -> dict[str, str]:
    return {
        "role": message.role,
        "content": message.content,
        "timestamp": message.timestamp.isoformat(timespec="seconds"),
    }


def _serialize_memory_context(context: AnswerContext | None) -> dict[str, Any]:
    if context is None:
        return {
            "used": False,
            "should_use": False,
            "score": 0,
            "threshold": 0,
            "reasons": [],
            "retrieval_modes": [],
            "reference_count": 0,
            "references": [],
            "retrieval_health": _default_retrieval_health(),
        }

    return {
        "used": context.used_memory,
        "should_use": context.should_use_memory,
        "score": context.score,
        "threshold": context.threshold,
        "reasons": list(context.reasons),
        "retrieval_modes": list(context.retrieval_modes),
        "reference_count": len(context.references),
        "references": [_serialize_memory_reference(reference) for reference in context.references],
        "retrieval_health": _serialize_retrieval_health(context),
    }


def _default_retrieval_health() -> dict[str, Any]:
    return {
        "index_status": "disabled",
        "index_reasons": [],
        "markdown_fallback_used": False,
        "retrieval_depth": "none",
        "query_variants": [],
        "core_enabled": False,
        "past_chats_enabled": False,
        "core_reference_count": 0,
        "structured_memory_hit_count": 0,
        "past_chat_hit_count": 0,
        "project_scope": None,
    }


def _serialize_retrieval_health(context: AnswerContext) -> dict[str, Any]:
    health = context.retrieval_health
    return {
        "index_status": _safe_log_text(str(health.index_status), max_length=40),
        "index_reasons": [
            _safe_log_text(str(reason), max_length=240)
            for reason in health.index_reasons[:10]
        ],
        "markdown_fallback_used": bool(health.markdown_fallback_used),
        "retrieval_depth": _safe_log_text(str(health.retrieval_depth), max_length=40),
        "query_variants": [
            _safe_log_text(str(variant), max_length=160)
            for variant in health.query_variants[:8]
        ],
        "core_enabled": bool(health.core_enabled),
        "past_chats_enabled": bool(health.past_chats_enabled),
        "core_reference_count": max(0, int(health.core_reference_count)),
        "structured_memory_hit_count": max(0, int(health.structured_memory_hit_count)),
        "past_chat_hit_count": max(0, int(health.past_chat_hit_count)),
        "project_scope": (
            _safe_log_text(str(health.project_scope), max_length=120)
            if health.project_scope is not None
            else None
        ),
    }


def _serialize_memory_reference(reference: MemoryContextReference) -> dict[str, Any]:
    return {
        "path": reference.path,
        "document_type": reference.document_type,
        "title": reference.title,
        "date": reference.date,
        "snippet": reference.snippet,
        "score": reference.score,
        "speaker_role": reference.speaker_role,
        "message_number": reference.message_number,
    }


def _normalize_attachments(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError("attachments must be a list.")
    if len(value) > MAX_ATTACHMENTS:
        raise ValueError(f"添付ファイルは1回の送信で最大{MAX_ATTACHMENTS}件までです。")

    reports = []
    for item in value:
        if not isinstance(item, dict):
            raise ValueError("attachment must be an object.")
        reports.append(_normalize_attachment(item))
    return reports


def _normalize_attachment(item: dict[str, Any]) -> dict[str, Any]:
    file_name = _safe_attachment_name(str(item.get("name") or item.get("file_name") or "attachment"))
    extension = Path(file_name).suffix.lower()
    size_bytes = _optional_int_value(item.get("size_bytes")) or 0
    report: dict[str, Any] = {
        "file_name": file_name,
        "extension": extension.lstrip(".") or "unknown",
        "size_bytes": size_bytes,
        "status": "error",
        "error": None,
        "extracted_chars": 0,
        "truncated": bool(item.get("truncated", False)),
        "_text": "",
    }

    if extension not in ALLOWED_ATTACHMENT_EXTENSIONS:
        report["error"] = "未対応のファイル形式です。"
        return report
    if size_bytes > MAX_ATTACHMENT_BYTES:
        report["error"] = f"ファイルサイズが上限({MAX_ATTACHMENT_BYTES} bytes)を超えています。"
        return report

    extractor_truncated = False
    try:
        if extension in TEXT_ATTACHMENT_EXTENSIONS:
            text = _attachment_text(item)
        elif extension == ".pdf":
            text = _extract_pdf_text_from_attachment(item)
        elif extension == ".xlsx":
            text, extractor_truncated = _extract_xlsx_text_from_attachment(item)
        else:
            text = ""
    except Exception as exc:
        report["error"] = f"{type(exc).__name__}: {exc}"
        return report

    text = _normalize_attachment_text(text)
    truncated = bool(report["truncated"]) or extractor_truncated or len(text) > MAX_ATTACHMENT_TEXT_CHARS
    if truncated:
        text = text[:MAX_ATTACHMENT_TEXT_CHARS]

    if not text.strip():
        report["error"] = "抽出できる本文がありません。"
        return report

    report.update(
        {
            "status": "extracted",
            "error": None,
            "extracted_chars": len(text),
            "truncated": truncated,
            "_text": text,
        }
    )
    return report


def _attachment_text(item: dict[str, Any]) -> str:
    text = item.get("text")
    if isinstance(text, str):
        return text

    raw = _attachment_bytes(item)
    return raw.decode("utf-8", errors="replace")


def _attachment_bytes(item: dict[str, Any]) -> bytes:
    data_base64 = item.get("data_base64")
    if not isinstance(data_base64, str) or not data_base64:
        raise ValueError("添付ファイル本文がありません。")
    try:
        data = base64.b64decode(data_base64, validate=True)
    except binascii.Error as exc:
        raise ValueError("添付ファイルのbase64が不正です。") from exc
    if len(data) > MAX_ATTACHMENT_BYTES:
        raise ValueError(f"ファイルサイズが上限({MAX_ATTACHMENT_BYTES} bytes)を超えています。")
    return data


def _extract_pdf_text_from_attachment(item: dict[str, Any]) -> str:
    raw = _attachment_bytes(item)
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF抽出には pypdf が必要です。") from exc

    import io

    reader = PdfReader(io.BytesIO(raw))
    texts = []
    for page in reader.pages[:20]:
        texts.append(page.extract_text() or "")
    return "\n\n".join(texts)


def _extract_xlsx_text_from_attachment(item: dict[str, Any]) -> tuple[str, bool]:
    raw = _attachment_bytes(item)
    workbook = _load_xlsx_workbook(raw)
    lines: list[str] = []
    extracted_length = 0
    truncated = False

    try:
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if len(visible_sheets) > MAX_XLSX_SHEETS:
            truncated = True

        for sheet in visible_sheets[:MAX_XLSX_SHEETS]:
            max_row = min(sheet.max_row or 1, MAX_XLSX_ROWS_PER_SHEET)
            max_column = min(sheet.max_column or 1, MAX_XLSX_COLUMNS_PER_SHEET)
            if (sheet.max_row or 0) > max_row or (sheet.max_column or 0) > max_column:
                truncated = True

            sheet_lines: list[str] = []
            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                    values_only=True,
                ),
                start=1,
            ):
                values = [_xlsx_cell_text(value) for value in row]
                while values and not values[-1]:
                    values.pop()
                if not values:
                    continue

                row_text = f"Row {row_number}: " + "\t".join(values)
                remaining = MAX_ATTACHMENT_TEXT_CHARS + 1 - extracted_length
                if len(row_text) > remaining:
                    row_text = row_text[:remaining]
                    truncated = True
                sheet_lines.append(row_text)
                extracted_length += len(row_text) + 1
                if extracted_length > MAX_ATTACHMENT_TEXT_CHARS:
                    truncated = True
                    break

            if sheet_lines:
                header = f"# Sheet: {_xlsx_cell_text(sheet.title)}"
                lines.extend([header, *sheet_lines, ""])
                extracted_length += len(header) + 2
            if extracted_length > MAX_ATTACHMENT_TEXT_CHARS:
                break
    finally:
        workbook.close()

    return "\n".join(lines).strip(), truncated


def _load_xlsx_workbook(raw: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel抽出には openpyxl が必要です。") from exc

    import io

    try:
        return load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        raise ValueError(f"Excelファイルを開けませんでした: {exc}") from exc


def _xlsx_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, datetime_time)):
        return value.isoformat()
    return re.sub(r"[\r\n\t]+", " ", str(value)).strip()


def _normalize_attachment_text(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def _safe_attachment_name(value: str) -> str:
    name = Path(value.replace("\\", "/")).name.strip() or "attachment"
    name = re.sub(r"[\x00-\x1f<>:\"/\\|?*]+", "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    if not name:
        name = "attachment"
    if len(name) > 120:
        suffix = Path(name).suffix
        stem_limit = max(1, 120 - len(suffix) - 3)
        name = f"{Path(name).stem[:stem_limit]}...{suffix}"
    return name


def _format_attachment_metadata_for_saved_message(attachments: list[dict[str, Any]]) -> str:
    if not attachments:
        return ""

    lines = ["", "", "[Attachments]"]
    for attachment in attachments:
        parts = [
            f"name={attachment['file_name']}",
            f"type={attachment['extension']}",
            f"status={attachment['status']}",
            f"chars={attachment['extracted_chars']}",
            f"truncated={'yes' if attachment['truncated'] else 'no'}",
        ]
        if attachment.get("error"):
            parts.append(f"error={_safe_log_text(str(attachment['error']), max_length=160)}")
        lines.append("- " + "; ".join(parts))
    return "\n".join(lines)


def _format_user_content_for_generation(content: str, attachments: list[dict[str, Any]]) -> str:
    context = _attachment_context_text(attachments)
    if not context:
        return content
    return "\n\n".join(
        [
            content,
            "---",
            "添付ファイルの一時コンテキストです。本文はlive JSONLには保存せず、この回答生成だけに使います。",
            context,
        ]
    )


def _attachment_context_text(attachments: list[dict[str, Any]]) -> str:
    sections = []
    for attachment in attachments:
        if attachment.get("status") != "extracted":
            continue
        sections.extend(
            [
                f"## {attachment['file_name']}",
                f"type: {attachment['extension']}, chars: {attachment['extracted_chars']}, truncated: {attachment['truncated']}",
                "",
                str(attachment.get("_text") or ""),
            ]
        )
    return "\n".join(sections).strip()


def _messages_for_generation(
    messages: list[LiveMessage],
    saved_user_message: LiveMessage,
    prompt_user_content: str,
) -> list[LiveMessage]:
    if prompt_user_content == saved_user_message.content:
        return messages

    generation_messages = list(messages)
    generation_messages[-1] = LiveMessage(
        role=saved_user_message.role,
        content=prompt_user_content,
        timestamp=saved_user_message.timestamp,
    )
    return generation_messages


def _serialize_attachment_report(attachment: dict[str, Any]) -> dict[str, Any]:
    return {
        "file_name": attachment["file_name"],
        "extension": attachment["extension"],
        "size_bytes": attachment["size_bytes"],
        "status": attachment["status"],
        "error": attachment["error"],
        "extracted_chars": attachment["extracted_chars"],
        "truncated": attachment["truncated"],
    }


def _optional_int_value(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _serialize_saved_session(session: Any, root: Path) -> dict[str, Any]:
    return {
        "session_id": session.session_id,
        "status": session.status,
        "title": session.title,
        "jsonl_file": _display_path(session.jsonl_file, root),
        "metadata_file": _display_path(session.metadata_file, root),
        "message_count": session.message_count,
        "started_at": session.started_at.isoformat(timespec="seconds"),
        "updated_at": session.updated_at.isoformat(timespec="seconds"),
        "saved_at": session.saved_at.isoformat(timespec="seconds"),
    }


def _serialize_resume_session(session: Any, root: Path) -> dict[str, Any]:
    return {
        "session_id": session.session_id,
        "title": session.title,
        "jsonl_file": _display_path(session.jsonl_file, root),
        "message_count": session.message_count,
        "started_at": session.started_at.isoformat(timespec="seconds"),
        "updated_at": session.updated_at.isoformat(timespec="seconds"),
        "last_user_at": session.last_user_at.isoformat(timespec="seconds"),
        "organization": _serialize_organization(root=root, session_file=session.jsonl_file),
        "personalization": serialize_personalization(load_session_personalization(root, session.jsonl_file)),
    }


def _serialize_organization(root: Path, session_file: Path) -> dict[str, Any]:
    if not session_file.exists():
        organization = {
            "status": "empty",
            "label": "未開始",
            "can_organize": False,
            "is_organized": False,
            "next_stage": None,
            "failed_stage": None,
            "last_error": None,
            "raw_file": None,
            "task_file": None,
            "current_message_count": 0,
            "current_updated_at": None,
            "organized_message_count": 0,
            "organized_updated_at": None,
            "stages": {
                "raw": {"name": "raw", "label": "raw.md作成", "status": "pending"},
                "memory": {"name": "memory", "label": "記憶整理", "status": "pending"},
                "index": {"name": "index", "label": "検索index更新", "status": "pending"},
            },
        }
    else:
        organization = get_session_organization(root=root, session_file=session_file)

    if load_session_personalization(root, session_file).exclude_from_memory:
        organization = {
            **organization,
            "status": "temporary",
            "label": "一時チャット",
            "can_organize": False,
            "is_organized": False,
            "next_stage": None,
            "failed_stage": None,
            "last_error": None,
        }
    return organization


def _display_path(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return str(path)


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


if __name__ == "__main__":
    raise SystemExit(main())
