import argparse
import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook


def safe_name(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    return name.strip() or "sheet"


def cell_value_with_merged(sheet, row, col):
    """
    結合セルなら左上セルの値を返す。
    AI用には結合セルを展開しておいた方が読みやすい。
    """
    cell = sheet.cell(row=row, column=col)

    if cell.value is not None:
        return cell.value

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = sheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )
            return top_left.value

    return None


def trim_table(rows):
    """
    末尾の空行・空列を削る。
    """
    # 空行削除
    while rows and all(v is None or v == "" for v in rows[-1]):
        rows.pop()

    if not rows:
        return []

    # 最大列数をそろえる
    max_cols = max(len(r) for r in rows)
    rows = [r + [None] * (max_cols - len(r)) for r in rows]

    # 空列削除
    while rows and rows[0] and all(
        row[-1] is None or row[-1] == "" for row in rows
    ):
        for row in rows:
            row.pop()

    return rows


def normalize_value(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def rows_to_markdown(rows, max_rows=80):
    if not rows:
        return "_No data_"

    rows = rows[:max_rows]
    header = [str(normalize_value(v)) for v in rows[0]]
    body = rows[1:]

    md = []
    md.append("| " + " | ".join(header) + " |")
    md.append("| " + " | ".join(["---"] * len(header)) + " |")

    for row in body:
        vals = [str(normalize_value(v)).replace("\n", " ") for v in row]
        md.append("| " + " | ".join(vals) + " |")

    return "\n".join(md)


def guess_headers(rows):
    if not rows:
        return []

    first = rows[0]
    headers = []
    for i, v in enumerate(first, start=1):
        text = str(normalize_value(v))
        headers.append(text if text else f"Column_{i}")
    return headers


def rows_to_records(rows):
    if len(rows) < 2:
        return []

    headers = guess_headers(rows)
    records = []

    for row in rows[1:]:
        record = {}
        for h, v in zip(headers, row):
            record[h] = normalize_value(v)
        records.append(record)

    return records


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel workbook into AI-readable CSV, Markdown, and JSON files."
    )
    parser.add_argument("input", help="Path to .xlsx file")
    parser.add_argument("--out", default="ai_readable", help="Output directory")
    parser.add_argument("--include-hidden", action="store_true", help="Include hidden sheets")
    args = parser.parse_args()

    input_path = Path(args.input)
    out_dir = Path(args.out)
    sheets_dir = out_dir / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)

    # data_only=True: 数式セルはExcel側で保存済みの計算結果を読む
    wb = load_workbook(input_path, data_only=True)

    summary_lines = []
    summary_lines.append(f"# Workbook Summary")
    summary_lines.append("")
    summary_lines.append(f"- File: `{input_path.name}`")
    summary_lines.append(f"- Sheets: {len(wb.sheetnames)}")
    summary_lines.append("")

    workbook_json = {
        "file": input_path.name,
        "sheets": []
    }

    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible" and not args.include_hidden:
            continue

        rows = []
        for r in range(1, sheet.max_row + 1):
            row_values = []
            for c in range(1, sheet.max_column + 1):
                row_values.append(
                    normalize_value(cell_value_with_merged(sheet, r, c))
                )
            rows.append(row_values)

        rows = trim_table(rows)
        filename = safe_name(sheet.title)

        csv_path = sheets_dir / f"{filename}.csv"
        md_path = sheets_dir / f"{filename}.md"
        json_path = sheets_dir / f"{filename}.json"

        # CSV
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        # Markdown
        markdown = rows_to_markdown(rows)
        md_path.write_text(
            f"# Sheet: {sheet.title}\n\n{markdown}\n",
            encoding="utf-8"
        )

        # JSON
        records = rows_to_records(rows)
        json_path.write_text(
            json.dumps(records, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8"
        )

        row_count = len(rows)
        col_count = max((len(r) for r in rows), default=0)
        headers = guess_headers(rows)

        summary_lines.append(f"## {sheet.title}")
        summary_lines.append("")
        summary_lines.append(f"- Size: {row_count} rows × {col_count} columns")
        summary_lines.append(f"- CSV: `sheets/{filename}.csv`")
        summary_lines.append(f"- Markdown: `sheets/{filename}.md`")
        summary_lines.append(f"- JSON: `sheets/{filename}.json`")
        summary_lines.append(f"- Headers: {', '.join(headers[:20])}")
        summary_lines.append("")

        workbook_json["sheets"].append({
            "name": sheet.title,
            "rows": row_count,
            "columns": col_count,
            "headers": headers,
            "files": {
                "csv": str(csv_path),
                "markdown": str(md_path),
                "json": str(json_path)
            }
        })

    (out_dir / "workbook_summary.md").write_text(
        "\n".join(summary_lines),
        encoding="utf-8"
    )

    (out_dir / "workbook_summary.json").write_text(
        json.dumps(workbook_json, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"Done: {out_dir}")
    print("AIに読ませるなら、まず workbook_summary.md を渡すといいです。")


if __name__ == "__main__":
    main()